from flask import Flask, request, Response, render_template, jsonify, send_from_directory
import json, os, re, subprocess, math, zipfile, shutil, sys
from urllib.parse import quote
from datetime import datetime
from collections import OrderedDict
import numpy as np

APP_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(APP_DIR, os.pardir, os.pardir))

def load_local_env():
    env_path = os.path.join(PROJECT_ROOT, '.env')
    if not os.path.isfile(env_path):
        return
    with open(env_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value

load_local_env()

DATA_DIR = os.environ.get('SENSECV_DATA_DIR', os.path.join(PROJECT_ROOT, 'data'))
CLIPS_DIR = os.environ.get('SENSECV_CLIPS_DIR', os.path.join(DATA_DIR, 'clips'))
DATASETS_DIR = os.environ.get('SENSECV_DATASETS_DIR', os.path.join(DATA_DIR, 'datasets'))
EXPORTS_DIR = os.environ.get('SENSECV_EXPORTS_DIR', os.path.join(DATA_DIR, 'exports'))
HISTORY_FILE = os.environ.get('SENSECV_HISTORY_FILE', os.path.join(DATA_DIR, 'history.json'))
UPLOADS_DIR = os.environ.get('SENSECV_UPLOADS_DIR', os.path.join(DATA_DIR, 'uploaded_datasets'))
DERIVED_DIR = os.environ.get('SENSECV_DERIVED_DIR', os.path.join(DATA_DIR, 'derived'))
MODELS_DIR = os.environ.get('SENSECV_MODELS_DIR', os.path.join(DATA_DIR, 'models'))
os.makedirs(CLIPS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(DERIVED_DIR, exist_ok=True)
os.makedirs(MODELS_DIR, exist_ok=True)

app = Flask(__name__, template_folder=os.path.join(APP_DIR, 'templates'))
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('SENSECV_MAX_UPLOAD_BYTES', str(2 * 1024 * 1024 * 1024)))

@app.route('/health')
def health():
    refresh_clips()
    return jsonify({'status': 'ok', 'clips': len(CLIPS)})

@app.after_request
def no_cache(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ─── Clips ───────────────────────────────────────────────────────────────────

# Extra input roots to scan for clip subfolders, in addition to CLIPS_DIR.
# Keep machine-specific dataset paths in SENSECV_EXTRA_CLIP_ROOTS instead of code.
EXTRA_CLIP_ROOTS = [DATASETS_DIR]
ENV_EXTRA_CLIP_ROOTS = [
    p.strip() for p in (
        os.environ.get('SENSECV_EXTRA_CLIP_ROOTS')
        or ''
    ).split(os.pathsep)
    if p.strip()
]
EXTRA_CLIP_ROOTS.extend(ENV_EXTRA_CLIP_ROOTS + [UPLOADS_DIR])

# Maps clip display name -> absolute folder path. Populated by find_clips().
CLIP_PATHS = {}
CLIP_VIDEO_PATHS = {}
CLIP_GROUPS = {}
# Display names that are filmed horizontally the whole time (the SenseCV roots).
# For these we skip vertical-orientation detection and only segment walking.
WALKING_ONLY = set()

def _clip_video_file(p):
    video = os.path.join(p, 'video.mp4')
    if os.path.isfile(video):
        return video
    if not os.path.isdir(p):
        return None
    mp4s = sorted(
        os.path.join(p, name) for name in os.listdir(p)
        if name.lower().endswith('.mp4') and os.path.isfile(os.path.join(p, name))
    )
    return mp4s[0] if mp4s else None

def _is_clip_dir(p):
    return os.path.isdir(p) and _clip_video_file(p) is not None \
        and os.path.isfile(os.path.join(p,'frames.json'))

def _safe_name(name):
    name = re.sub(r'[^\w.\-]+', '_', os.path.splitext(os.path.basename(name))[0]).strip('._')
    return name or 'dataset'

def _unique_dir(parent, name):
    base = _safe_name(name)
    target = os.path.join(parent, base)
    if not os.path.exists(target):
        return target
    i = 2
    while True:
        candidate = os.path.join(parent, f'{base}_{i}')
        if not os.path.exists(candidate):
            return candidate
        i += 1

def _safe_extract_zip(zip_file, target_dir):
    os.makedirs(target_dir, exist_ok=True)
    for info in zip_file.infolist():
        normalized = info.filename.replace('\\', '/')
        parts = [p for p in normalized.split('/') if p]
        if not parts or normalized.startswith('/') or any(p == '..' for p in parts):
            raise ValueError(f'unsafe zip path: {info.filename}')
        out_path = os.path.abspath(os.path.join(target_dir, *parts))
        target_abs = os.path.abspath(target_dir)
        if os.path.commonpath([target_abs, out_path]) != target_abs:
            raise ValueError(f'unsafe zip path: {info.filename}')
        if info.is_dir():
            os.makedirs(out_path, exist_ok=True)
            continue
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with zip_file.open(info) as src, open(out_path, 'wb') as dst:
            shutil.copyfileobj(src, dst)

def _discover_clip_dirs(root, recursive=False):
    if not os.path.isdir(root):
        return []
    found = []
    if not recursive:
        for name in sorted(os.listdir(root)):
            p = os.path.join(root, name)
            if _is_clip_dir(p):
                found.append((name, p, None))
        return found
    for current, dirs, _files in os.walk(root):
        if _is_clip_dir(current):
            rel = os.path.relpath(current, root).replace(os.sep, '/')
            first = rel.split('/', 1)[0] if rel != '.' else os.path.basename(os.path.normpath(root))
            found.append((rel, current, first))
            dirs[:] = []
    return sorted(found, key=lambda item: item[0])

def find_clips():
    """Scan CLIPS_DIR plus any EXTRA_CLIP_ROOTS for clip subfolders.

    Clips in the primary root keep their bare folder name (so existing history
    and exports stay valid); clips from extra roots are prefixed with the root
    folder name to keep display names unique across roots.

    The shared lookup dicts are built locally and swapped in atomically at the
    end. With threaded=True, a concurrent reader (e.g. the long inspection loop
    while the UI polls /api/export-state) must never observe a half-cleared
    dict; rebinding the module globals is atomic, clear()+repopulate is not.
    """
    global CLIP_PATHS, CLIP_VIDEO_PATHS, CLIP_GROUPS, WALKING_ONLY
    clip_paths, clip_video_paths, clip_groups, walking_only_set = {}, {}, {}, set()
    clips = []
    # (root, label, walking_only). The primary root is the supermarket footage
    # (held vertical); the extra roots are filmed horizontally throughout.
    recursive_roots = {
        os.path.normcase(os.path.abspath(r))
        for r in [DATASETS_DIR] + ENV_EXTRA_CLIP_ROOTS + [UPLOADS_DIR]
    }
    roots = []
    seen_roots = set()

    def add_root(root, label, walking_only, recursive):
        root_abs = os.path.normcase(os.path.abspath(root))
        if root_abs in seen_roots:
            return
        seen_roots.add(root_abs)
        roots.append((root, label, walking_only, recursive))

    # When CLIPS_DIR coincides with a recursive root (e.g. SENSECV_DATA_DIR
    # deployments where /data/clips is also the uploads root), it must keep
    # the recursive scan or nested uploaded datasets are never found.
    if os.path.normcase(os.path.abspath(CLIPS_DIR)) in recursive_roots:
        add_root(CLIPS_DIR, None, True, True)
    else:
        add_root(CLIPS_DIR, None, False, False)
    for r in EXTRA_CLIP_ROOTS:
        recursive = os.path.normcase(os.path.abspath(r)) in recursive_roots
        label = None if recursive else os.path.basename(os.path.normpath(r))
        add_root(r, label, True, recursive)
    if os.path.isdir(EXPORTS_DIR):
        add_root(EXPORTS_DIR, 'exports', True, False)
    for root, label, walking_only, recursive in roots:
        if not os.path.isdir(root):
            continue
        for display_name, clip_dir, discovered_group in _discover_clip_dirs(root, recursive):
            display = display_name if label is None else f"{label}/{display_name}"
            group = discovered_group if recursive and discovered_group else label or 'Supermercado Telefrango'
            clip_paths[display] = clip_dir
            clip_video_paths[display] = _clip_video_file(clip_dir)
            clip_groups[display] = group
            if walking_only:
                walking_only_set.add(display)
            clips.append(display)
    # Atomic swap: readers see either the old maps or the fully-built new ones.
    CLIP_PATHS, CLIP_VIDEO_PATHS, CLIP_GROUPS, WALKING_ONLY = (
        clip_paths, clip_video_paths, clip_groups, walking_only_set)
    return clips

CLIPS = find_clips()

def refresh_clips():
    global CLIPS
    CLIPS = find_clips()
    return CLIPS

def clip_path(idx):
    return CLIP_PATHS[CLIPS[idx]]

def clip_video_path(idx):
    return CLIP_VIDEO_PATHS[CLIPS[idx]]

def clip_groups():
    return [CLIP_GROUPS.get(name, '') for name in CLIPS]

# ─── History ─────────────────────────────────────────────────────────────────

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8-sig') as f:
            history = json.load(f)
        pruned = prune_history_to_exports(history)
        if len(pruned) != len(history):
            save_history(pruned)
        return pruned
    return []

def save_history(h):
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(h, f, indent=2, ensure_ascii=False)

def list_export_folders():
    if not os.path.isdir(EXPORTS_DIR):
        return []
    return sorted(
        name for name in os.listdir(EXPORTS_DIR)
        if os.path.isdir(os.path.join(EXPORTS_DIR, name))
    )

def prune_history_to_exports(history):
    export_folders = set(list_export_folders())
    return [entry for entry in history if entry.get('folder') in export_folders]

def get_next_number():
    used = set()
    for name in list_export_folders():
        m = re.match(r'^(\d+)(?:_|$)', name)
        if m:
            used.add(int(m.group(1)))
    number = 1
    while number in used:
        number += 1
    return number

def name_exists(folder_name):
    return os.path.exists(os.path.join(EXPORTS_DIR, folder_name))

def learned_walking_window(idx):
    """Return the latest exported walking crop for this source clip, if any."""
    source = CLIPS[idx]
    for entry in reversed(load_history()):
        if entry.get('source_idx') != idx and entry.get('source_clip') != source:
            continue
        start = entry.get('start')
        end = entry.get('end')
        if start is None or end is None or end <= start:
            continue
        return round(float(start), 3), round(float(end), 3)
    return None

# ─── Data processing ─────────────────────────────────────────────────────────

def load_json(folder, name):
    with open(os.path.join(folder, name), 'r', encoding='utf-8-sig') as f:
        return json.load(f)

def interp_at_times(src_t, src_v, query_t):
    n = len(src_t); result = []
    for qt in query_t:
        idx = int(np.searchsorted(src_t, qt))
        if idx <= 0:            result.append(src_v[0].tolist())
        elif idx >= n:          result.append(src_v[-1].tolist())
        else:
            t0,t1 = src_t[idx-1],src_t[idx]; v0,v1 = src_v[idx-1],src_v[idx]
            f = (qt-t0)/(t1-t0) if t1!=t0 else 0.
            result.append((v0+f*(v1-v0)).tolist())
    return result

def external_input_at_times(folder, query_t):
    path = os.path.join(folder, 'external_sensors.json')
    if not os.path.isfile(path):
        return []
    samples = load_json(folder, 'external_sensors.json').get('external_sensors', [])
    if not samples:
        return []
    sensor_t = np.array([s.get('time_usec', 0) for s in samples], dtype=np.float64)
    buttons = np.array([1 if int(s.get('button', s.get('input', 0)) or 0) == 1 else 0
                        for s in samples], dtype=np.int8)
    result = []
    for qt in query_t:
        idx = int(np.searchsorted(sensor_t, qt, side='right')) - 1
        if idx < 0:
            result.append(0)
        elif idx >= len(buttons):
            result.append(int(buttons[-1]))
        else:
            result.append(int(buttons[idx]))
    return result

# Weinberg step-length constant: L = K * (a_max - a_min)^(1/4) per step, with
# acceleration in m/s². ~0.5 gives realistic ~0.6-0.9 m strides; tune per gait.
STEP_LENGTH_K = float(os.environ.get('SENSECV_STEP_LENGTH_K', '0.5'))


def _step_forward_speed(acc_t_us, acc_v, query_us, k=STEP_LENGTH_K):
    """Forward speed (m/s) from gait *steps*, with no acceleration integration.

    Pedestrian dead-reckoning instead of double-integration (which drifts):
      1. Dynamic acceleration = |a| minus its ~0.5 s rolling mean (drops gravity;
         the magnitude is orientation-independent and pulses once per step).
      2. Detect steps as peaks above an adaptive threshold, honouring a
         refractory gap (cadence capped near 3.3 Hz).
      3. Each step's length is Weinberg's L = k·(a_max − a_min)^(1/4) over the
         step interval; forward speed across that interval = L / step_period.
      4. Speed is held per step interval and 0 where there is no stepping.
    Returned aligned to `query_us` (sample-and-hold).
    """
    t = np.asarray(acc_t_us, dtype=np.float64) / 1e6
    q = np.asarray(query_us, dtype=np.float64) / 1e6
    if len(t) < 8:
        return np.zeros(len(q))
    amag = np.linalg.norm(np.asarray(acc_v, dtype=np.float64), axis=1)
    dt = float(np.median(np.diff(t)))
    if not (dt > 1e-4):
        dt = 0.01
    fs = 1.0 / dt
    win = max(3, int(round(0.5 * fs)))
    base = np.convolve(amag, np.ones(win) / win, mode='same')
    dyn = amag - base
    thr = max(0.6 * float(np.std(dyn)), 0.4)
    min_gap = max(1, int(round(0.30 * fs)))     # cadence ceiling ~3.3 Hz
    peaks = []
    last = -(10 ** 9)
    for i in range(1, len(dyn) - 1):
        if (dyn[i] > thr and dyn[i] >= dyn[i - 1] and dyn[i] > dyn[i + 1]
                and (i - last) >= min_gap):
            peaks.append(i)
            last = i
    # Per-step speed (Weinberg length / step period), placed at the step midpoint.
    step_t, step_v = [], []
    for j in range(1, len(peaks)):
        a, b = peaks[j - 1], peaks[j]
        period = t[b] - t[a]
        if not (0.30 <= period <= 1.25):        # plausible step period (0.8-3.3 Hz)
            continue
        seg = amag[a:b + 1]
        rng = float(seg.max() - seg.min())
        if rng <= 0:
            continue
        step_t.append(0.5 * (t[a] + t[b]))
        step_v.append((k * rng ** 0.25) / period)
    if not step_t:
        return np.zeros(len(q))

    # Build a *continuous* speed curve (not a stair): interpolate between step
    # midpoints, ramping up from rest at the clip start and back down to rest if
    # stepping stops well before the end. Then smooth lightly.
    anchor_t = [t[0]] + step_t
    anchor_v = [0.0] + step_v
    if (t[-1] - step_t[-1]) > 0.6:              # stepping ceased -> ramp to rest
        anchor_t += [step_t[-1] + 0.3, t[-1]]
        anchor_v += [0.0, 0.0]
    else:
        anchor_t += [t[-1]]
        anchor_v += [step_v[-1]]
    cont = np.interp(t, anchor_t, anchor_v)
    sm = max(3, int(round(0.4 * fs)))
    cont = np.convolve(cont, np.ones(sm) / sm, mode='same')
    qi = np.clip(np.searchsorted(t, q), 0, len(t) - 1)
    return cont[qi]


def _heading_lateral_axis(vel_xy):
    """Forward/lateral unit axes in the horizontal plane from movement direction.

    Forward = dominant horizontal-velocity direction (speed-weighted principal
    component, signed so the mean motion is positive). Lateral = forward rotated
    +90° (left of travel). Returns (forward_hat, lateral_hat) as 2-vectors.
    """
    H = np.asarray(vel_xy, dtype=np.float64).reshape(-1, 2)
    spd = np.linalg.norm(H, axis=1)
    if not spd.any():
        return np.array([1.0, 0.0]), np.array([0.0, 1.0])
    M = (H * spd[:, None]).T @ H                 # speed-weighted scatter
    _w, vecs = np.linalg.eigh(M)
    fwd = vecs[:, -1]                            # principal direction
    mean_v = (H * spd[:, None]).sum(0)
    if float(np.dot(mean_v, fwd)) < 0:
        fwd = -fwd
    n = float(np.linalg.norm(fwd))
    fwd = fwd / n if n > 1e-9 else np.array([1.0, 0.0])
    lat = np.array([-fwd[1], fwd[0]])
    return fwd, lat


def compute_velocities(acc_data, rot_data, ft_us):
    acc_t = np.array([a['time_usec'] for a in acc_data], dtype=np.float64)
    acc_v = np.array([[a['x'],a['y'],a['z']] for a in acc_data])
    rot_t = np.array([r['time_usec'] for r in rot_data], dtype=np.float64)
    rot_v = np.array([[r['x'],r['y'],r['z']] for r in rot_data])
    ft    = np.array(ft_us, dtype=np.float64)

    n = min(200, len(acc_v)); g_phone = acc_v[:n].mean(0)
    G = float(np.linalg.norm(g_phone)); G = G if G>1e-3 else 9.81
    g_hat = g_phone/G; g_world = np.array([0.,0.,1.])
    cross = np.cross(g_hat, g_world); dot = float(np.dot(g_hat, g_world)); cn = float(np.linalg.norm(cross))
    if cn<1e-7: q = np.array([1.,0.,0.,0.]) if dot>0 else np.array([0.,1.,0.,0.])
    else:
        s = np.sqrt(max(0.,(1.+dot)*2.)); q = np.array([s/2,cross[0]/s,cross[1]/s,cross[2]/s]); q/=np.linalg.norm(q)

    def qmul(a,b):
        w1,x1,y1,z1=a; w2,x2,y2,z2=b
        return np.array([w1*w2-x1*x2-y1*y2-z1*z2, w1*x2+x1*w2+y1*z2-z1*y2,
                         w1*y2-x1*z2+y1*w2+z1*x2, w1*z2+x1*y2-y1*x2+z1*w2])
    def qrot(q,v):
        w,x,y,z=q
        return np.array([
            (1-2*(y*y+z*z))*v[0]+2*(x*y-z*w)*v[1]+2*(x*z+y*w)*v[2],
            2*(x*y+z*w)*v[0]+(1-2*(x*x+z*z))*v[1]+2*(y*z-x*w)*v[2],
            2*(x*z-y*w)*v[0]+2*(y*z+x*w)*v[1]+(1-2*(x*x+y*y))*v[2],
        ])

    vel=np.zeros(3); rec=[(rot_t[0],vel.copy())]; prev_t=rot_t[0]
    for i in range(1,len(rot_t)):
        t=rot_t[i]; dt=(t-prev_t)/1e6; prev_t=t
        if dt<=0 or dt>0.05: rec.append((t,vel.copy())); continue
        om=rot_v[i]; on=float(np.linalg.norm(om))
        if on>1e-8:
            ha=on*dt/2; ax=om/on
            dq=np.array([np.cos(ha),np.sin(ha)*ax[0],np.sin(ha)*ax[1],np.sin(ha)*ax[2]])
            q=qmul(q,dq); q/=np.linalg.norm(q)
        idx=min(max(int(np.searchsorted(acc_t,t)),0),len(acc_v)-1)
        aw=qrot(q,acc_v[idx]); an=aw/max(float(np.linalg.norm(aw)),1e-6)
        err=np.cross(an,np.array([0.,0.,1.])); cor=np.array([1.,err[0]*.005,err[1]*.005,err[2]*.005])
        q=qmul(q,cor/np.linalg.norm(cor)); q/=np.linalg.norm(q)
        aw=qrot(q,acc_v[idx]); al=aw-np.array([0.,0.,G])
        al=np.where(np.abs(al)<0.25,0.,al); vel+=al*dt; vel*=0.998
        rec.append((t,vel.copy()))

    vt=np.array([r[0] for r in rec]); vv=np.array([r[1] for r in rec])
    # Lateral velocity = leaky-integrated horizontal velocity projected onto the
    # axis perpendicular to the travel heading (the pipeline lateral component).
    _fwd, lat_hat = _heading_lateral_axis(vv[:, :2])
    lat_at_rec = vv[:, :2] @ lat_hat
    # Forward velocity from gait steps (drift-free), sampled at frame times.
    fwd_at_frame = _step_forward_speed(acc_t, acc_v, ft)
    result=[]
    for fi, f in enumerate(ft):
        i=min(max(int(np.searchsorted(vt,f)),0),len(vv)-1); v=vv[i]
        result.append({'vx':float(v[0]),'vy':float(v[1]),'vz':float(v[2]),
                       'speed':float(np.linalg.norm(v)),
                       'lateral':float(lat_at_rec[i]),
                       'forward':float(fwd_at_frame[fi])})
    return result

_cache = OrderedDict()
_walking_template_cache = {'mtime': None, 'templates': None, 'durations': None}
_walking_classifier_cache = {'mtime': None, 'model': None}
_classifier_feature_cache = OrderedDict()
_dronet_runtime = {'model': None, 'preprocess_bgr': None, 'torch': None, 'error': None}
_dronet_cache = OrderedDict()
_ssim_selection_cache = OrderedDict()
DRONET_DIR = os.environ.get('DRONET_DIR', os.path.join(PROJECT_ROOT, "third_party", "dronet"))
DRONET_WEIGHTS = os.environ.get('DRONET_WEIGHTS', os.path.join(DRONET_DIR, "repo", "model", "model_weights.h5"))
DRONET_SAMPLE_FPS = 3.0
SSIM_THRESHOLD = float(os.environ.get('SENSECV_SSIM_THRESHOLD', '0.985'))
SSIM_MAX_GAP_SEC = float(os.environ.get('SENSECV_SSIM_MAX_GAP_SEC', '0.5'))
SSIM_REVIEW_FPS = float(os.environ.get('SENSECV_SSIM_REVIEW_FPS', '15'))
SSIM_REVIEW_MAX_WIDTH = int(os.environ.get('SENSECV_SSIM_REVIEW_MAX_WIDTH', '960'))

def _video_rotation(idx):
    """Display-rotation tag (degrees CW) the browser will honor for this clip's
    video, read from the container metadata. The viewer counter-rotates by this
    so square/tagged clips are not shown on their side. Returns 0 if unknown."""
    const = None
    try:
        import cv2
        const = getattr(cv2, 'CAP_PROP_ORIENTATION_META', None)
        if const is None:
            return 0
        cap = cv2.VideoCapture(clip_video_path(idx))
        if not cap.isOpened():
            return 0
        rot = cap.get(const)
        cap.release()
        return int(round(rot)) % 360
    except Exception:
        return 0

def get_clip_data(idx):
    key = CLIPS[idx]
    if key in _cache: _cache.move_to_end(key); return _cache[key]
    folder = clip_path(idx)
    frames   = load_json(folder,'frames.json')['frames']
    acc_data = load_json(folder,'accelerations.json')['accelerations']
    rot_data = load_json(folder,'rotations.json')['rotations']
    ft = np.array([f['time_usec'] for f in frames], dtype=np.float64)
    at = np.array([a['time_usec'] for a in acc_data], dtype=np.float64)
    av = np.array([[a['x'],a['y'],a['z']] for a in acc_data])
    rt = np.array([r['time_usec'] for r in rot_data], dtype=np.float64)
    rv = np.array([[r['x'],r['y'],r['z']] for r in rot_data])
    t0=float(ft[0]); secs=[(float(f)-t0)/1e6 for f in ft]
    fps=float(1e6/np.median(np.diff(ft)))
    data={'fps':round(fps,3),'duration':secs[-1],'times':secs,
          'accel':interp_at_times(at,av,ft),'rotation':interp_at_times(rt,rv,ft),
          'velocity':compute_velocities(acc_data,rot_data,ft.tolist()),
          'external_input':external_input_at_times(folder, ft),
          'name':key,'index':idx,'total':len(CLIPS),
          'video_rotation':_video_rotation(idx)}
    _cache[key]=data
    if len(_cache)>5: _cache.popitem(last=False)
    return data

def _load_dronet_runtime():
    if _dronet_runtime['model'] is not None or _dronet_runtime['error'] is not None:
        return _dronet_runtime
    try:
        if DRONET_DIR not in sys.path:
            sys.path.insert(0, DRONET_DIR)
        import torch
        from dronet_model import load_dronet, preprocess_bgr
        _dronet_runtime['model'] = load_dronet(DRONET_WEIGHTS)
        _dronet_runtime['preprocess_bgr'] = preprocess_bgr
        _dronet_runtime['torch'] = torch
    except Exception as e:
        _dronet_runtime['error'] = str(e)
    return _dronet_runtime

def _dronet_direction(steering):
    if abs(steering) < 0.1:
        return 'STRAIGHT'
    return 'RIGHT' if steering > 0 else 'LEFT'

def dronet_frame_classification(idx, time_s, exact=False):
    runtime = _load_dronet_runtime()
    if runtime['error']:
        return {'available': False, 'error': runtime['error']}

    video_path = clip_video_path(idx)
    mtime = os.path.getmtime(video_path)
    try:
        cap = cv2.VideoCapture(video_path)
    except NameError:
        import cv2 as _cv2
        globals()['cv2'] = _cv2
        cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        return {'available': False, 'error': 'video unreadable'}

    fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    if fps <= 0 or frame_count <= 0:
        cap.release()
        return {'available': False, 'error': 'invalid video metadata'}

    duration = frame_count / fps
    sample_time = max(0.0, min(float(time_s or 0.0), duration))
    if not exact:
        sample_time = math.floor(sample_time * DRONET_SAMPLE_FPS) / DRONET_SAMPLE_FPS
    frame_idx = min(frame_count - 1, max(0, int(round(sample_time * fps))))

    key = (CLIPS[idx], mtime, frame_idx)
    if key in _dronet_cache:
        _dronet_cache.move_to_end(key)
        cap.release()
        return _dronet_cache[key]

    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
    ok, frame = cap.read()
    cap.release()
    if not ok:
        return {'available': False, 'error': 'frame decode failed', 'frame': frame_idx}

    tensor, _crop = runtime['preprocess_bgr'](frame)
    with runtime['torch'].no_grad():
        steering_t, collision_t = runtime['model'](tensor)
    steering = float(steering_t.item())
    collision = float(collision_t.item())
    yaw = steering * 90.0
    result = {
        'available': True,
        'clip': CLIPS[idx],
        'frame': frame_idx,
        'time_s': round(frame_idx / fps, 4),
        'requested_time_s': round(float(time_s or 0.0), 4),
        'sample_fps': DRONET_SAMPLE_FPS,
        'exact': bool(exact),
        'source_fps': round(fps, 4),
        'steering': steering,
        'yaw_deg': yaw,
        'direction': _dronet_direction(steering),
        'collision_prob': collision,
        'collision_label': 'COLLISION' if collision >= 0.5 else 'CLEAR',
    }
    _dronet_cache[key] = result
    if len(_dronet_cache) > 300:
        _dronet_cache.popitem(last=False)
    return result

# ─── Sensor data save ────────────────────────────────────────────────────────

def save_sensor_data(clip_idx, start_sec, end_sec, export_folder):
    folder = clip_path(clip_idx)
    frames_raw = load_json(folder,'frames.json')
    acc_raw    = load_json(folder,'accelerations.json')
    rot_raw    = load_json(folder,'rotations.json')
    t0_usec = frames_raw['frames'][0]['time_usec']
    t_start = t0_usec + start_sec * 1e6
    t_end   = t0_usec + end_sec   * 1e6
    # re-base time_usec so t_start → 0
    def rebase(t): return t - t_start

    filtered_frames = [
        {**f, 'time_usec': int(rebase(f['time_usec'])),
               'sensor_timestamp': int(rebase(f['sensor_timestamp']))}
        for f in frames_raw['frames']
        if t_start <= f['time_usec'] <= t_end
    ]
    # reset frame_id
    for i, fr in enumerate(filtered_frames): fr['frame_id'] = i

    filtered_acc = [
        {**a, 'time_usec': int(rebase(a['time_usec']))}
        for a in acc_raw['accelerations'] if t_start <= a['time_usec'] <= t_end
    ]
    filtered_rot = [
        {**r, 'time_usec': int(rebase(r['time_usec']))}
        for r in rot_raw['rotations'] if t_start <= r['time_usec'] <= t_end
    ]
    external_raw = None
    external_path = os.path.join(folder, 'external_sensors.json')
    if os.path.isfile(external_path):
        external_raw = load_json(folder, 'external_sensors.json')
    filtered_external = []
    if external_raw:
        filtered_external = [
            {**s, 'time_usec': int(rebase(s['time_usec']))}
            for s in external_raw.get('external_sensors', [])
            if t_start <= s.get('time_usec', -1) <= t_end
        ]
    with open(os.path.join(export_folder,'frames.json'),'w') as f:
        json.dump({'frames': filtered_frames}, f)
    with open(os.path.join(export_folder,'accelerations.json'),'w') as f:
        json.dump({'accelerations': filtered_acc}, f)
    with open(os.path.join(export_folder,'rotations.json'),'w') as f:
        json.dump({'rotations': filtered_rot}, f)
    if external_raw is not None:
        with open(os.path.join(export_folder,'external_sensors.json'),'w') as f:
            json.dump({'external_sensors': filtered_external}, f)

# ─── Suggest crop ────────────────────────────────────────────────────────────

def _frame_records_in_window(clip_idx, start_sec, end_sec):
    frames = load_json(clip_path(clip_idx), 'frames.json')['frames']
    if not frames:
        return []
    t0 = float(frames[0]['time_usec'])
    records = []
    for source_index, frame in enumerate(frames):
        time_s = (float(frame['time_usec']) - t0) / 1e6
        if start_sec <= time_s <= end_sec:
            records.append((source_index, frame, time_s))
    return records


def _ssim_gray(a, b):
    a = a.astype(np.float32)
    b = b.astype(np.float32)
    c1 = (0.01 * 255.0) ** 2
    c2 = (0.03 * 255.0) ** 2
    mu_a = float(a.mean())
    mu_b = float(b.mean())
    da = a - mu_a
    db = b - mu_b
    var_a = float((da * da).mean())
    var_b = float((db * db).mean())
    cov = float((da * db).mean())
    denom = (mu_a * mu_a + mu_b * mu_b + c1) * (var_a + var_b + c2)
    if denom <= 1e-12:
        return 1.0 if np.array_equal(a, b) else 0.0
    return ((2 * mu_a * mu_b + c1) * (2 * cov + c2)) / denom


def _decode_gray_frame(cap, frame_no):
    try:
        import cv2
    except Exception as e:
        raise RuntimeError(f'OpenCV indisponivel para SSIM: {e}')
    cap.set(cv2.CAP_PROP_POS_FRAMES, int(frame_no))
    ok, frame = cap.read()
    if not ok:
        return None
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    return cv2.resize(gray, (160, 90), interpolation=cv2.INTER_AREA)


SELECTION_METRICS = ('ssim', 'dinov2', 'lpips', 'vif')


class _Dinov2Backend:
    """Cosine similarity between DINOv2 ViT-S/14 CLS embeddings."""

    def __init__(self):
        import torch
        self.torch = torch
        self.model = torch.hub.load('facebookresearch/dinov2', 'dinov2_vits14')
        self.model.eval()
        self.mean = torch.tensor([0.485, 0.456, 0.406]).view(3, 1, 1)
        self.std = torch.tensor([0.229, 0.224, 0.225]).view(3, 1, 1)

    def prepare(self, frame_bgr):
        import cv2
        rgb = cv2.cvtColor(cv2.resize(frame_bgr, (224, 224), interpolation=cv2.INTER_AREA),
                           cv2.COLOR_BGR2RGB)
        tensor = self.torch.from_numpy(rgb).permute(2, 0, 1).float() / 255.0
        tensor = (tensor - self.mean) / self.std
        with self.torch.no_grad():
            emb = self.model(tensor.unsqueeze(0))[0]
        return emb / emb.norm()

    def similarity(self, prev, cur):
        return float((prev * cur).sum())


class _LpipsBackend:
    """1 - LPIPS(AlexNet) perceptual distance."""

    def __init__(self):
        import torch
        import lpips
        self.torch = torch
        self.net = lpips.LPIPS(net='alex', verbose=False)
        self.net.eval()

    def prepare(self, frame_bgr):
        import cv2
        rgb = cv2.cvtColor(cv2.resize(frame_bgr, (256, 144), interpolation=cv2.INTER_AREA),
                           cv2.COLOR_BGR2RGB)
        tensor = self.torch.from_numpy(rgb).permute(2, 0, 1).float() / 255.0
        return (tensor * 2.0 - 1.0).unsqueeze(0)

    def similarity(self, prev, cur):
        with self.torch.no_grad():
            return 1.0 - float(self.net(prev, cur))


class _VifBackend:
    """Pixel-domain Visual Information Fidelity on the SSIM-sized grayscale."""

    def __init__(self):
        from sewar.full_ref import vifp
        self.vifp = vifp

    def prepare(self, frame_bgr):
        import cv2
        gray = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2GRAY)
        return cv2.resize(gray, (160, 90), interpolation=cv2.INTER_AREA)

    def similarity(self, prev, cur):
        return float(self.vifp(prev, cur))


_metric_backend_classes = {'dinov2': _Dinov2Backend, 'lpips': _LpipsBackend, 'vif': _VifBackend}
_metric_backends = {}


def _metric_backend(metric):
    if metric not in _metric_backends:
        _metric_backends[metric] = _metric_backend_classes[metric]()
    return _metric_backends[metric]


def _decode_color_frame(cap, frame_no):
    try:
        import cv2
    except Exception as e:
        raise RuntimeError(f'OpenCV indisponivel para selecao de frames: {e}')
    cap.set(cv2.CAP_PROP_POS_FRAMES, int(frame_no))
    ok, frame = cap.read()
    return frame if ok else None


def ssim_frame_selection(clip_idx, start_sec, end_sec,
                         threshold=SSIM_THRESHOLD,
                         max_gap_sec=SSIM_MAX_GAP_SEC,
                         metric='ssim'):
    """Select visually distinct frames inside a proposed crop."""
    records = _frame_records_in_window(clip_idx, start_sec, end_sec)
    before = len(records)
    empty = {
        'frames_before': before,
        'frames_after': 0,
        'metric': metric,
        'ssim_threshold': threshold,
        'ssim_max_gap_sec': max_gap_sec,
        'selected_frames': [],
    }
    if not records:
        return empty

    video_path = clip_video_path(clip_idx)
    key = (
        CLIPS[clip_idx],
        os.path.getmtime(video_path),
        round(float(start_sec), 3),
        round(float(end_sec), 3),
        float(threshold),
        float(max_gap_sec),
        metric,
    )
    if key in _ssim_selection_cache:
        _ssim_selection_cache.move_to_end(key)
        return _ssim_selection_cache[key]

    try:
        import cv2
        cap = cv2.VideoCapture(video_path)
    except Exception as e:
        return {**empty, 'error': f'OpenCV indisponivel para SSIM: {e}'}
    if not cap.isOpened():
        return {**empty, 'error': 'video unreadable for SSIM'}

    video_fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
    if video_fps <= 0:
        cap.release()
        return {**empty, 'error': 'invalid video fps for SSIM'}

    try:
        backend = None if metric == 'ssim' else _metric_backend(metric)
    except Exception as e:
        cap.release()
        return {**empty, 'error': f'metric {metric} indisponivel: {e}'}

    selected = []
    prev_rep = None
    prev_time = None
    last_record_i = len(records) - 1
    for record_i, (source_index, frame, time_s) in enumerate(records):
        frame_no = max(0, int(round(time_s * video_fps)))
        if backend is None:
            rep = _decode_gray_frame(cap, frame_no)
        else:
            color = _decode_color_frame(cap, frame_no)
            rep = None if color is None else backend.prepare(color)
        if rep is None:
            continue

        if prev_rep is None:
            score = None
        elif backend is None:
            score = _ssim_gray(prev_rep, rep)
        else:
            score = backend.similarity(prev_rep, rep)
        forced_gap = prev_time is not None and (time_s - prev_time) >= max_gap_sec
        keep = (
            prev_rep is None
            or record_i == last_record_i
            or score is None
            or score < threshold
            or forced_gap
        )
        if keep:
            selected.append({
                'source_index': int(source_index),
                'frame_id': int(frame.get('frame_id', source_index)),
                'time_s': round(float(time_s), 3),
                'ssim_prev': None if score is None else round(float(score), 5),
            })
            prev_rep = rep
            prev_time = time_s
    cap.release()

    result = {
        'frames_before': before,
        'frames_after': len(selected),
        'metric': metric,
        'ssim_threshold': threshold,
        'ssim_max_gap_sec': max_gap_sec,
        'selected_frames': selected,
    }
    _ssim_selection_cache[key] = result
    if len(_ssim_selection_cache) > 64:
        _ssim_selection_cache.popitem(last=False)
    return result


DEFAULT_METRIC_THRESHOLDS = {
    'ssim': SSIM_THRESHOLD,
    'dinov2': 0.98,
    'lpips': 0.95,
    'vif': 0.70,
}


def _coerce_metric(value):
    metric = (value or 'ssim').strip().lower()
    if metric not in SELECTION_METRICS:
        raise ValueError(f"metric must be one of: {', '.join(SELECTION_METRICS)}")
    return metric


def _coerce_ssim_threshold(value, metric='ssim'):
    if value is None or value == '':
        return DEFAULT_METRIC_THRESHOLDS.get(metric, SSIM_THRESHOLD)
    threshold = float(value)
    if not 0.0 < threshold < 1.0:
        raise ValueError('SSIM threshold must be between 0 and 1')
    return threshold


def _request_metric():
    return _coerce_metric(request.args.get('metric'))


def _request_ssim_threshold(metric='ssim'):
    return _coerce_ssim_threshold(request.args.get('threshold'), metric=metric)


def _body_ssim_threshold(body, metric='ssim'):
    return _coerce_ssim_threshold((body or {}).get('ssim_threshold'), metric=metric)


def _selection_public(selection):
    return {
        k: v for k, v in selection.items()
        if k != 'selected_frames'
    }


def _suggestion_with_ssim(idx, suggestion, threshold=SSIM_THRESHOLD, metric='ssim'):
    if not suggestion.get('found'):
        return suggestion
    try:
        selection = ssim_frame_selection(
            idx,
            float(suggestion['start']),
            float(suggestion['end']),
            threshold=threshold,
            metric=metric,
        )
    except Exception as e:
        selection = {'frames_before': 0, 'frames_after': 0, 'error': str(e)}
    enriched = dict(suggestion)
    enriched['ssim'] = _selection_public(selection)
    return enriched


def save_ssim_selection(selection, export_folder):
    with open(os.path.join(export_folder, 'ssim_selection.json'), 'w', encoding='utf-8') as f:
        json.dump(selection, f, indent=2)


def _review_frame_size(width, height, max_width=SSIM_REVIEW_MAX_WIDTH):
    width = max(1, int(width or 640))
    height = max(1, int(height or 360))
    max_width = max(1, int(max_width or width))
    if width <= max_width:
        out_w, out_h = width, height
    else:
        ratio = max_width / float(width)
        out_w, out_h = max_width, max(1, int(round(height * ratio)))
    return max(2, out_w - (out_w % 2)), max(2, out_h - (out_h % 2))


def _write_ssim_review_video(src_video, frame_nos, out_video, fps=SSIM_REVIEW_FPS):
    try:
        import cv2
    except Exception as e:
        raise RuntimeError(f'OpenCV indisponivel para video SSIM: {e}')

    cap = cv2.VideoCapture(src_video)
    if not cap.isOpened():
        raise RuntimeError('video unreadable for SSIM review')

    width, height = _review_frame_size(
        cap.get(cv2.CAP_PROP_FRAME_WIDTH),
        cap.get(cv2.CAP_PROP_FRAME_HEIGHT),
    )
    try:
        import imageio_ffmpeg
        ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    except Exception as e:
        cap.release()
        raise RuntimeError(f'ffmpeg indisponivel para video SSIM: {e}')

    cmd = [
        ffmpeg, '-y',
        '-f', 'rawvideo',
        '-vcodec', 'rawvideo',
        '-pix_fmt', 'rgb24',
        '-s', f'{width}x{height}',
        '-r', f'{max(1.0, float(fps)):.3f}',
        '-i', '-',
        '-an',
        '-c:v', 'libx264',
        '-preset', 'ultrafast',
        '-crf', '23',
        '-pix_fmt', 'yuv420p',
        '-movflags', '+faststart',
        out_video,
    ]
    proc = subprocess.Popen(cmd, stdin=subprocess.PIPE,
                            stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    written = 0
    try:
        for frame_no in frame_nos:
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(frame_no))
            ok, frame = cap.read()
            if not ok:
                continue
            if frame.shape[1] != width or frame.shape[0] != height:
                frame = cv2.resize(frame, (width, height), interpolation=cv2.INTER_AREA)
            proc.stdin.write(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB).tobytes())
            written += 1

        if written == 0:
            empty = np.zeros((height, width, 3), dtype=np.uint8)
            cv2.putText(empty, 'No frames', (24, max(42, height // 2)),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.0, (220, 220, 220), 2, cv2.LINE_AA)
            proc.stdin.write(cv2.cvtColor(empty, cv2.COLOR_BGR2RGB).tobytes())
    finally:
        cap.release()
        if proc.stdin:
            proc.stdin.close()
            proc.stdin = None
    _stdout, stderr = proc.communicate()
    if proc.returncode != 0:
        raise RuntimeError((stderr or b'').decode('utf-8', errors='ignore')[-500:])
    return written


def save_ssim_review_videos(clip_idx, start_sec, end_sec, selection, export_folder):
    """Write audit videos for all candidate, selected, and rejected SSIM frames."""
    records = _frame_records_in_window(clip_idx, start_sec, end_sec)
    video_path = clip_video_path(clip_idx)
    try:
        import cv2
    except Exception as e:
        raise RuntimeError(f'OpenCV indisponivel para video SSIM: {e}')

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise RuntimeError('video unreadable for SSIM review')
    video_fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
    cap.release()
    if video_fps <= 0:
        raise RuntimeError('invalid video fps for SSIM review')

    selected_sources = {
        int(frame['source_index'])
        for frame in selection.get('selected_frames', [])
        if 'source_index' in frame
    }
    all_items = [
        (source_index, max(0, int(round(time_s * video_fps))))
        for source_index, _frame, time_s in records
    ]
    chosen = [frame_no for source_index, frame_no in all_items
              if int(source_index) in selected_sources]
    not_chosen = [frame_no for source_index, frame_no in all_items
                  if int(source_index) not in selected_sources]

    review_dir = os.path.join(export_folder, 'ssim_review')
    os.makedirs(review_dir, exist_ok=True)
    counts = {
        'all_frames': _write_ssim_review_video(
            video_path, [frame_no for _source_index, frame_no in all_items],
            os.path.join(review_dir, 'all_frames.mp4'),
        ),
        'chosen_frames': _write_ssim_review_video(
            video_path, chosen, os.path.join(review_dir, 'chosen_frames.mp4'),
        ),
        'not_chosen_frames': _write_ssim_review_video(
            video_path, not_chosen, os.path.join(review_dir, 'not_chosen_frames.mp4'),
        ),
    }
    with open(os.path.join(review_dir, 'manifest.json'), 'w', encoding='utf-8') as f:
        json.dump({
            'source_video': video_path,
            'start_sec': round(float(start_sec), 3),
            'end_sec': round(float(end_sec), 3),
            'review_fps': SSIM_REVIEW_FPS,
            'counts': counts,
        }, f, indent=2)
    return counts


def ssim_selection_payload(clip_idx, start_sec, end_sec, threshold=SSIM_THRESHOLD, metric='ssim'):
    selection = ssim_frame_selection(clip_idx, start_sec, end_sec, threshold=threshold, metric=metric)
    payload = dict(selection)
    frames = []
    for frame in payload.get('selected_frames', []):
        item = dict(frame)
        item['image_url'] = f"/frame/{clip_idx}/{int(item['source_index'])}.jpg"
        frames.append(item)
    payload['selected_frames'] = frames
    return payload


def _orientation_walking_masks(idx):
    """
    Per-frame boolean masks for phone orientation and gait.

    vertical (portrait):
      ay/|a| > 0.95  →  gravity nearly along phone y-axis (portrait)
      az/|a| < 0.12  →  screen not facing up/down (camera pointing forward)
      Both smoothed over a 1 s window.
      Calibrated against 5 clips with known ground-truth start times
      (clip1→13.0s, clip2→11.5s, clip3→none, clip4→9.0s, clip5→none): 5/5.

    walking:
      During a walk the gravity-magnitude |a| oscillates with each step,
      so the rolling standard deviation of |a| over ~1 s rises well above
      the near-static value. std > 0.5 m/s² flags sustained gait.
    """
    folder = clip_path(idx)
    frames   = load_json(folder, 'frames.json')['frames']
    acc_data = load_json(folder, 'accelerations.json')['accelerations']

    ft = np.array([f['time_usec'] for f in frames], dtype=np.float64)
    at = np.array([a['time_usec'] for a in acc_data], dtype=np.float64)
    av = np.array([[a['x'], a['y'], a['z']] for a in acc_data])
    if len(ft) < 2 or len(av) == 0:
        raise ValueError('sensor data is empty or too short')

    t0 = float(ft[0])
    secs = (ft - t0) / 1e6
    fps  = float(1e6 / np.median(np.diff(ft)))

    amag = np.empty(len(ft))
    ay_n = np.empty(len(ft))
    az_n = np.empty(len(ft))
    for i, a_usec in enumerate(ft):
        j = min(max(int(np.searchsorted(at, a_usec)), 0), len(av)-1)
        a = av[j]
        m = math.sqrt(a[0]**2 + a[1]**2 + a[2]**2)
        amag[i] = m
        ay_n[i] = abs(a[1]) / m if m > 1 else 0.
        az_n[i] = abs(a[2]) / m if m > 1 else 0.

    # Smooth / window over 1 s. Keep the kernel no longer than the clip; numpy
    # returns the larger length for mode='same' when the kernel is longer.
    W = min(len(ft), max(1, int(fps * 1.0)))
    kernel = np.ones(W) / W
    s_ay = np.convolve(ay_n, kernel, mode='same')
    s_az = np.convolve(az_n, kernel, mode='same')
    vertical = (s_ay > 0.95) & (s_az < 0.12)

    # Rolling std of |a| over the same window → gait energy
    mean_mag = np.convolve(amag, kernel, mode='same')
    var_mag  = np.convolve((amag - mean_mag) ** 2, kernel, mode='same')
    std_mag  = np.sqrt(np.maximum(var_mag, 0.))
    walking  = std_mag > 0.5

    return secs, fps, vertical, walking


def _imu_walking_series(idx):
    """Per-frame portrait and walking evidence from accelerometer + gyroscope."""
    folder = clip_path(idx)
    frames = load_json(folder, 'frames.json')['frames']
    acc_data = load_json(folder, 'accelerations.json')['accelerations']
    rot_data = load_json(folder, 'rotations.json')['rotations']

    ft = np.array([f['time_usec'] for f in frames], dtype=np.float64)
    at = np.array([a['time_usec'] for a in acc_data], dtype=np.float64)
    av = np.array([[a['x'], a['y'], a['z']] for a in acc_data], dtype=np.float64)
    rt = np.array([r['time_usec'] for r in rot_data], dtype=np.float64)
    rv = np.array([[r['x'], r['y'], r['z']] for r in rot_data], dtype=np.float64)
    if len(ft) < 2 or len(av) == 0 or len(rv) == 0:
        raise ValueError('sensor data is empty or too short')

    secs = (ft - float(ft[0])) / 1e6
    fps = float(1e6 / np.median(np.diff(ft)))

    aligned_acc = np.empty((len(ft), 3), dtype=np.float64)
    aligned_rot = np.empty((len(ft), 3), dtype=np.float64)
    for i, t in enumerate(ft):
        ai = min(max(int(np.searchsorted(at, t)), 0), len(av) - 1)
        ri = min(max(int(np.searchsorted(rt, t)), 0), len(rv) - 1)
        aligned_acc[i] = av[ai]
        aligned_rot[i] = rv[ri]

    amag = np.linalg.norm(aligned_acc, axis=1)
    gyro_mag = np.linalg.norm(aligned_rot, axis=1)
    ay_n = np.abs(aligned_acc[:, 1]) / np.maximum(amag, 1e-6)
    az_n = np.abs(aligned_acc[:, 2]) / np.maximum(amag, 1e-6)

    W = min(len(ft), max(1, int(fps * 0.8)))
    kernel = np.ones(W) / W
    s_ay = np.convolve(ay_n, kernel, mode='same')
    s_az = np.convolve(az_n, kernel, mode='same')

    trend = np.convolve(amag, kernel, mode='same')
    acc_hp = amag - trend
    acc_energy = np.sqrt(np.maximum(np.convolve(acc_hp * acc_hp, kernel, mode='same'), 0.))
    gyro_energy = np.convolve(gyro_mag, kernel, mode='same')
    jerk = np.convolve(np.abs(np.diff(amag, prepend=amag[0])) * fps, kernel, mode='same')

    def robust_z(values):
        med = float(np.median(values))
        iqr = float(np.percentile(values, 75) - np.percentile(values, 25))
        return (values - med) / max(iqr, 1e-3)

    walking_score = (
        1.20 * robust_z(acc_energy) +
        0.80 * robust_z(gyro_energy) +
        0.35 * robust_z(jerk)
    )
    portrait = (s_ay > 0.88) & (s_az < 0.30)
    strict_portrait = (s_ay > 0.95) & (s_az < 0.12)
    return secs, fps, portrait, strict_portrait, walking_score, acc_energy, gyro_energy


def _moving_average(values, frames):
    frames = min(len(values), max(1, int(frames)))
    return np.convolve(values, np.ones(frames) / frames, mode='same')


def _centered_rolling_mean(values, frames):
    """Centered moving average, same length as input.

    Name used by the IMU event detector ([[imu-event-labeling]]); identical to
    _moving_average but tolerant of list input.
    """
    return _moving_average(np.asarray(values, dtype=np.float64), frames)


def _rolling_std(values, frames):
    mean = _moving_average(values, frames)
    return np.sqrt(np.maximum(_moving_average((values - mean) ** 2, frames), 0.))


def _step_regularity(amag, fps, win_sec):
    """Rolling autocorrelation feature for gait-like step cadence."""
    n = len(amag)
    half = max(2, int(fps * win_sec / 2))
    lags = np.arange(max(2, int(fps * 0.28)), max(3, int(fps * 0.85)))
    regularity = np.zeros(n)
    cadence = np.zeros(n)
    power = np.zeros(n)
    if len(lags) == 0:
        return regularity, cadence, power

    min_window = int(lags[-1]) + 2
    for i in range(n):
        start = max(0, i - half)
        end = min(n, i + half + 1)
        window = amag[start:end]
        if len(window) < min_window:
            continue
        centered = window - np.mean(window)
        denom = float(np.dot(centered, centered))
        power[i] = math.sqrt(denom / len(centered)) if len(centered) else 0.
        if denom <= 1e-9:
            continue

        best_corr = -1.
        best_lag = 0
        for lag in lags:
            if lag >= len(centered):
                continue
            corr = float(np.dot(centered[:-lag], centered[lag:]) / (denom + 1e-9))
            if corr > best_corr:
                best_corr = corr
                best_lag = int(lag)
        regularity[i] = max(best_corr, 0.)
        cadence[i] = fps / best_lag if best_lag else 0.
    return regularity, cadence, power


def _classifier_feature_series(idx):
    """Frame features for the supervised walking/non-walking classifier."""
    key = CLIPS[idx]
    if key in _classifier_feature_cache:
        _classifier_feature_cache.move_to_end(key)
        return _classifier_feature_cache[key]

    folder = clip_path(idx)
    frames = load_json(folder, 'frames.json')['frames']
    acc_data = load_json(folder, 'accelerations.json')['accelerations']
    rot_data = load_json(folder, 'rotations.json')['rotations']

    ft = np.array([f['time_usec'] for f in frames], dtype=np.float64)
    at = np.array([a['time_usec'] for a in acc_data], dtype=np.float64)
    av = np.array([[a['x'], a['y'], a['z']] for a in acc_data], dtype=np.float64)
    rt = np.array([r['time_usec'] for r in rot_data], dtype=np.float64)
    rv = np.array([[r['x'], r['y'], r['z']] for r in rot_data], dtype=np.float64)
    if len(ft) < 2 or len(av) == 0 or len(rv) == 0:
        raise ValueError('sensor data is empty or too short')

    secs = (ft - float(ft[0])) / 1e6
    fps = float(1e6 / np.median(np.diff(ft)))

    aligned_acc = np.empty((len(ft), 3), dtype=np.float64)
    aligned_rot = np.empty((len(ft), 3), dtype=np.float64)
    for i, t in enumerate(ft):
        ai = min(max(int(np.searchsorted(at, t)), 0), len(av) - 1)
        ri = min(max(int(np.searchsorted(rt, t)), 0), len(rv) - 1)
        aligned_acc[i] = av[ai]
        aligned_rot[i] = rv[ri]

    amag = np.linalg.norm(aligned_acc, axis=1)
    gyro_mag = np.linalg.norm(aligned_rot, axis=1)
    ax_n = np.abs(aligned_acc[:, 0]) / np.maximum(amag, 1e-6)
    ay_n = np.abs(aligned_acc[:, 1]) / np.maximum(amag, 1e-6)
    az_n = np.abs(aligned_acc[:, 2]) / np.maximum(amag, 1e-6)

    features = []
    for sec in (0.35, 0.60, 0.90, 1.40):
        W = fps * sec
        s_ax = _moving_average(ax_n, W)
        s_ay = _moving_average(ay_n, W)
        s_az = _moving_average(az_n, W)
        trend = _moving_average(amag, W)
        acc_hp = amag - trend
        acc_energy = np.sqrt(np.maximum(_moving_average(acc_hp * acc_hp, W), 0.))
        gyro_energy = _moving_average(gyro_mag, W)
        gyro_std = _rolling_std(gyro_mag, W)
        jerk = _moving_average(np.abs(np.diff(amag, prepend=amag[0])) * fps, W)

        raw_energy = [
            np.log1p(acc_energy),
            np.log1p(gyro_energy),
            np.log1p(gyro_std),
            np.log1p(jerk),
        ]
        features.extend([s_ay, s_az, s_ax])
        features.extend(raw_energy)

        for values in raw_energy:
            med = float(np.median(values))
            iqr = float(np.percentile(values, 75) - np.percentile(values, 25))
            features.append((values - med) / max(iqr, 1e-3))

    for win_sec in (1.5, 2.0, 2.6):
        regularity, cadence, power = _step_regularity(amag, fps, win_sec)
        cadence_match = np.exp(-((cadence - 1.8) / 0.45) ** 2)
        features.extend([
            regularity,
            cadence,
            cadence_match,
            np.log1p(power),
            regularity * cadence_match,
        ])

    X = np.column_stack(features)
    portrait = (
        (_moving_average(ay_n, fps * 0.8) > 0.88) &
        (_moving_average(az_n, fps * 0.8) < 0.30)
    )
    result = (secs, fps, X, portrait)
    _classifier_feature_cache[key] = result
    if len(_classifier_feature_cache) > 32:
        _classifier_feature_cache.popitem(last=False)
    return result


def _first_sustained(mask, secs, fps, min_sec=1.5):
    """First run of ≥ min_sec frames where mask is True.
    Returns (start_sec, end_sec) where end is the last True frame anywhere
    after that run, or (None, None) if no qualifying run exists."""
    min_frames = int(fps * min_sec)
    count = 0
    start_i = None
    for i, v in enumerate(mask):
        if v:
            if count == 0:
                start_i = i
            count += 1
            if count >= min_frames:
                end_i = int(len(mask) - 1 - np.argmax(mask[::-1]))
                return round(float(secs[start_i]), 2), round(float(secs[end_i]), 2)
        else:
            count = 0
            start_i = None
    return None, None


def _runs(mask, fps, min_sec=1.0, max_gap_sec=0.7):
    min_frames = max(1, int(fps * min_sec))
    max_gap = max(0, int(fps * max_gap_sec))
    result = []
    start_i = None
    last_true_i = None
    gap = 0

    for i, value in enumerate(mask):
        if value:
            if start_i is None:
                start_i = i
            last_true_i = i
            gap = 0
        elif start_i is not None:
            gap += 1
            if gap > max_gap:
                if last_true_i - start_i + 1 >= min_frames:
                    result.append((start_i, last_true_i))
                start_i = None
                last_true_i = None
                gap = 0

    if start_i is not None and last_true_i - start_i + 1 >= min_frames:
        result.append((start_i, last_true_i))
    return result


def _walking_classifier_model():
    """Train/cache a frame classifier from exported walking windows."""
    mtime = os.path.getmtime(HISTORY_FILE) if os.path.exists(HISTORY_FILE) else None
    if (_walking_classifier_cache['mtime'] == mtime and
            _walking_classifier_cache['model'] is not None):
        return _walking_classifier_cache['model']

    try:
        from sklearn.ensemble import RandomForestClassifier
    except Exception:
        return None

    X_parts = []
    y_parts = []
    for entry in load_history():
        idx = entry.get('source_idx')
        start = entry.get('start')
        end = entry.get('end')
        if idx is None or start is None or end is None:
            continue
        if idx < 0 or idx >= len(CLIPS) or end <= start:
            continue
        try:
            secs, _fps, features, _portrait = _classifier_feature_series(int(idx))
        except Exception:
            continue
        labels = ((secs >= float(start)) & (secs <= float(end))).astype(int)
        if labels.any() and (~labels.astype(bool)).any():
            X_parts.append(features)
            y_parts.append(labels)

    if not X_parts:
        return None

    model = RandomForestClassifier(
        n_estimators=300,
        max_depth=9,
        min_samples_leaf=6,
        class_weight='balanced_subsample',
        random_state=7,
        n_jobs=1,
    )
    model.fit(np.vstack(X_parts), np.concatenate(y_parts))
    _walking_classifier_cache.update({'mtime': mtime, 'model': model})
    return model


def _classifier_walking_window(idx):
    """
    Classify each frame as walking/non-walking, then return one smoothed segment.

    Features combine orientation, accelerometer/gyroscope energy, and rolling
    step-cadence autocorrelation. The classifier is trained from the exported
    windows in history.json and refreshed after new exports.
    """
    model = _walking_classifier_model()
    if model is None:
        return _imu_walking_window(idx)

    secs, fps, features, portrait = _classifier_feature_series(idx)
    probabilities = model.predict_proba(features)[:, 1]
    probabilities = _moving_average(probabilities, fps * 0.6)

    active = portrait & (probabilities > 0.50)
    runs = _runs(active, fps, min_sec=0.9, max_gap_sec=0.65)
    if not runs:
        return None, None

    merged = []
    for start_i, end_i in runs:
        if merged and secs[start_i] - secs[merged[-1][1]] <= 1.0:
            merged[-1] = (merged[-1][0], end_i)
        else:
            merged.append((start_i, end_i))

    best = None
    for start_i, end_i in merged:
        duration = float(secs[end_i] - secs[start_i])
        if duration < 3.0:
            continue
        quality = (
            float(np.mean(probabilities[start_i:end_i + 1])) +
            0.08 * math.log(max(duration, 0.1)) -
            0.04 * abs(duration - 6.25)
        )
        if duration > 12.1:
            quality -= 1.0
        if best is None or quality > best[0]:
            best = (quality, start_i, end_i)

    if best is None:
        return None, None

    start_i, end_i = best[1], best[2]
    edge = portrait & (probabilities > 0.45)
    while start_i <= end_i and not edge[start_i]:
        start_i += 1
    while end_i >= start_i and not edge[end_i]:
        end_i -= 1
    if end_i <= start_i:
        return None, None

    max_duration = 12.1
    if secs[end_i] - secs[start_i] > max_duration:
        window = max(1, int(max_duration * fps))
        best_sub = None
        for sub_start in range(start_i, max(start_i + 1, end_i - window + 2)):
            sub_end = min(end_i, sub_start + window - 1)
            quality = float(np.mean(probabilities[sub_start:sub_end + 1]))
            if best_sub is None or quality > best_sub[0]:
                best_sub = (quality, sub_start, sub_end)
        start_i, end_i = best_sub[1], best_sub[2]

    if secs[end_i] - secs[start_i] < 3.0:
        return None, None

    return round(float(secs[start_i]), 2), round(float(secs[end_i]), 2)


def _imu_walking_window(idx):
    """
    Detect the single walking bout using IMU activity.

    Calibrated against the current five exported windows:
    - portrait gate removes handling/setup outside usable footage;
    - acceleration high-pass energy captures steps;
    - gyroscope energy catches body/phone sway during walking;
    - loose expansion recovers the start/end around the active core.
    """
    secs, fps, portrait, strict_portrait, score, acc_energy, gyro_energy = _imu_walking_series(idx)

    # If walking starts when the phone is raised into portrait, the step-energy
    # core can lag behind the true visual start. The ground truth includes this
    # portrait transition, so detect that case before trimming by activity.
    portrait_runs = _runs(strict_portrait, fps, min_sec=2.0, max_gap_sec=0.5)
    transition_candidates = []
    for start_i, end_i in portrait_runs:
        duration = float(secs[end_i] - secs[start_i])
        pre_start = max(0, start_i - int(fps * 2.0))
        pre_portrait = float(strict_portrait[pre_start:start_i].mean()) if start_i > pre_start else 0.
        run_acc = float(np.median(acc_energy[start_i:end_i + 1]))
        run_gyro = float(np.median(gyro_energy[start_i:end_i + 1]))
        if 10.0 <= duration <= 12.5 and pre_portrait < 0.40 and run_acc > 0.70 and run_gyro > 0.12:
            quality = duration + run_acc + run_gyro
            transition_candidates.append((quality, start_i, end_i))
    if transition_candidates:
        _quality, start_i, end_i = max(transition_candidates, key=lambda item: item[0])
        return round(float(secs[start_i]), 2), round(float(secs[end_i]), 2)

    active = (
        portrait &
        (score > -0.50) &
        (acc_energy > 0.15) &
        (gyro_energy > 0.03)
    )
    runs = _runs(active, fps, min_sec=1.0, max_gap_sec=0.8)
    if not runs:
        return None, None

    # Merge nearby active pieces inside the same walking bout.
    merged = []
    for start_i, end_i in runs:
        if merged and secs[start_i] - secs[merged[-1][1]] <= 1.5:
            merged[-1] = (merged[-1][0], end_i)
        else:
            merged.append((start_i, end_i))

    target_duration = 6.25  # median-ish duration from the 5 ground-truth exports
    best = None
    for start_i, end_i in merged:
        duration = float(secs[end_i] - secs[start_i])
        quality = (
            float(np.mean(score[start_i:end_i + 1])) +
            0.25 * math.log(max(duration, 0.1)) -
            0.06 * abs(duration - target_duration)
        )
        if best is None or quality > best[0]:
            best = (quality, start_i, end_i)

    start_i, end_i = best[1], best[2]
    loose = portrait & (score > -1.20) & (acc_energy > 0.12) & (gyro_energy > 0.02)
    while start_i > 0 and loose[start_i - 1]:
        start_i -= 1
    while end_i + 1 < len(secs) and loose[end_i + 1]:
        end_i += 1

    max_duration = 12.1  # 1.15x the longest current ground-truth export
    if secs[end_i] - secs[start_i] > max_duration:
        window = max(1, int(max_duration * fps))
        best_sub = None
        for sub_start in range(start_i, max(start_i + 1, end_i - window + 2)):
            sub_end = min(end_i, sub_start + window - 1)
            quality = float(np.mean(score[sub_start:sub_end + 1]))
            if best_sub is None or quality > best_sub[0]:
                best_sub = (quality, sub_start, sub_end)
        start_i, end_i = best_sub[1], best_sub[2]

    min_duration = 3.0  # rejects short bursts while keeping the shortest benchmark core
    if secs[end_i] - secs[start_i] < min_duration:
        return None, None

    return round(float(secs[start_i]), 2), round(float(secs[end_i]), 2)


def suggest_crop(idx, mode='walking'):
    """
    Suggest a walking crop window (first sustained portrait *and* walking period;
    horizontal SenseCV clips ignore orientation). `mode` is kept for call-site
    compatibility but only 'walking' is supported now — the old 'vertical' mode
    was retired. Always reports has_vertical so the UI can warn when a clip never
    goes vertical at all.
    """
    try:
        secs, fps, vertical, walking = _orientation_walking_masks(idx)
    except Exception as e:
        return {'found': False, 'mode': 'walking', 'has_vertical': False,
                'message': f'Dados do clipe invÃ¡lidos: {e}'}

    # Horizontal footage (SenseCV roots): orientation is irrelevant, so ignore
    # the vertical mask entirely and just segment the first sustained walking
    # period (reusing a previously exported window when one exists).
    if CLIPS[idx] in WALKING_ONLY:
        learned = learned_walking_window(idx)
        if learned:
            start, end = learned
            return {'found': True, 'mode': 'walking', 'start': start, 'end': end,
                    'has_vertical': True, 'learned': True}
        start, end = _first_sustained(walking, secs, fps)
        if start is not None:
            return {'found': True, 'mode': 'walking', 'start': start, 'end': end,
                    'has_vertical': True}
        return {'found': False, 'mode': 'walking', 'has_vertical': True,
                'message': 'Nenhum momento de caminhada sustentado encontrado'}

    has_vertical = bool(vertical.any())

    learned = learned_walking_window(idx)
    if learned:
        start, end = learned
        return {'found': True, 'mode': 'walking', 'start': start, 'end': end,
                'has_vertical': has_vertical, 'learned': True}
    if not has_vertical:
        start, end = None, None
    else:
        start, end = _classifier_walking_window(idx)
    if start is not None:
        return {'found': True, 'mode': 'walking', 'start': start, 'end': end,
                'has_vertical': has_vertical}

    if not has_vertical:
        msg = 'Nenhum período em posição vertical encontrado'
    else:
        msg = 'Posição vertical encontrada, mas sem momento de caminhada sustentado'
    return {'found': False, 'mode': 'walking', 'has_vertical': has_vertical, 'message': msg}

# ─── Lateral-deviation detection ─────────────────────────────────────────────

def suggest_lateral_deviation(idx, window_sec=1.0, dir_window_sec=2.0):
    """Window where the videomaker's lateral velocity is highest — sidestep.

    Calibrated against the operator's manually-cut deviation exports in
    `history.json` (entries 6-10, SenseCV clips 28-31 and 46). Of the
    features tried in `_debug_lateral.py`, **mean |lateral velocity| over a
    sliding 1 s window** matched the GT windows best (mean abs offset 0.61 s,
    beating yaw rate at 0.68 s and lateral-accel magnitude entirely).

    Algorithm:
      1. Gravity = 1 s rolling mean of accel → orientation-independent
         horizontal plane.
      2. Horizontal acceleration = `linear − (linear · g_hat) g_hat`,
         projected onto an orthonormal 2-D basis (e1, e2) in that plane.
      3. Integrate to a 2-D velocity, drift-corrected by subtracting the
         clip-mean (removes IMU bias accumulated by `cumsum`).
      4. Local walking direction = `dir_window_sec` rolling mean of that
         velocity. "Lateral" velocity = the component perpendicular to it.
      5. Slide a `window_sec`-wide window and pick the position with the
         highest mean `|lateral velocity|`.

    Always returns a window — per the user's "must happen in every lateral
    deviation clip" requirement.
    """
    folder = clip_path(idx)
    try:
        frames = load_json(folder, 'frames.json')['frames']
        acc    = load_json(folder, 'accelerations.json')['accelerations']
    except Exception as e:
        return {'found': False, 'mode': 'lateral', 'has_vertical': True,
                'message': f'Dados do clipe invÃ¡lidos: {e}'}
    ft = np.array([f['time_usec'] for f in frames], dtype=np.float64)
    if len(ft) < 2 or len(acc) == 0:
        return {'found': False, 'mode': 'lateral', 'has_vertical': True,
                'message': 'Dados do clipe incompletos'}
    at = np.array([a['time_usec'] for a in acc], dtype=np.float64)
    av = np.array([[a['x'], a['y'], a['z']] for a in acc])
    secs = (ft - ft[0]) / 1e6
    fps  = float(1e6 / np.median(np.diff(ft)))

    a_at_f = np.array(interp_at_times(at, av, ft.tolist()))

    # Gravity & horizontal acceleration
    Wg = min(len(ft), max(1, int(fps * 1.0)))
    kg = np.ones(Wg) / Wg
    g_vec = np.column_stack([np.convolve(a_at_f[:, k], kg, mode='same')
                             for k in range(3)])
    g_hat = g_vec / np.maximum(np.linalg.norm(g_vec, axis=1, keepdims=True), 1e-6)
    horiz3 = (a_at_f - g_vec) - np.sum((a_at_f - g_vec) * g_hat,
                                       axis=1, keepdims=True) * g_hat

    # 2-D basis in horizontal plane
    e1 = np.cross(g_hat, np.array([1., 0., 0.]))
    bad = np.linalg.norm(e1, axis=1) < 1e-3
    if bad.any():
        e1[bad] = np.cross(g_hat[bad], np.array([0., 1., 0.]))
    e1 /= np.maximum(np.linalg.norm(e1, axis=1, keepdims=True), 1e-6)
    e2 = np.cross(g_hat, e1)
    e2 /= np.maximum(np.linalg.norm(e2, axis=1, keepdims=True), 1e-6)
    horiz2 = np.column_stack([np.sum(horiz3 * e1, axis=1),
                              np.sum(horiz3 * e2, axis=1)])

    # Integrate to velocity, remove drift bias
    dt = np.diff(np.concatenate([[secs[0]], secs]))
    vel2 = np.cumsum(horiz2 * dt[:, None], axis=0)
    vel2 -= vel2.mean(axis=0, keepdims=True)

    # Local walking direction → lateral component
    Wd = min(len(ft), max(1, int(fps * dir_window_sec)))
    kd = np.ones(Wd) / Wd
    mean_v = np.column_stack([np.convolve(vel2[:, 0], kd, mode='same'),
                              np.convolve(vel2[:, 1], kd, mode='same')])
    mn  = np.maximum(np.linalg.norm(mean_v, axis=1, keepdims=True), 1e-6)
    fwd = mean_v / mn
    lat = np.column_stack([-fwd[:, 1], fwd[:, 0]])
    lat_vel = np.abs(np.sum(vel2 * lat, axis=1))

    w = max(1, int(round(fps * window_sec)))
    if len(lat_vel) <= w:
        return {'found': True, 'mode': 'lateral', 'has_vertical': True,
                'start': round(float(secs[0]), 3),
                'end':   round(float(secs[-1]), 3)}
    sliding_mean = np.convolve(lat_vel, np.ones(w), mode='valid') / w
    i = int(np.argmax(sliding_mean))
    return {'found': True, 'mode': 'lateral', 'has_vertical': True,
            'start': round(float(secs[i]), 3),
            'end':   round(float(secs[i + w - 1]), 3)}


# ─── Batch export ────────────────────────────────────────────────────────────

# ─── IMU criteria: capture validation + event labeling ──────────────────────
# Implements docs/wiki reference "criterios_imu_video_orientando" (uploaded
# PDF): validate IMU capture quality, detect physical events (desvio,
# reducao, parada) with T1/T2 from the IMU, and emit the three label windows
# (acao = T1..T2, decisao = T1-Δ..T1, expandido = T1-Δ..T2+margem) plus a
# confidence level (alta / baixa / descartar).

IMU_EVENT_DELTA_SEC = float(os.environ.get('SENSECV_IMU_EVENT_DELTA_SEC', '1.0'))
IMU_EVENT_PERSIST_SEC = float(os.environ.get('SENSECV_IMU_EVENT_PERSIST_SEC', '0.4'))
IMU_STOP_MIN_SEC = float(os.environ.get('SENSECV_IMU_STOP_MIN_SEC', '1.0'))
# Stop-onset fallback for clips with no desvio: a standstill stretch shorter
# than IMU_STOP_MIN_SEC still marks where the person started to halt, because
# recordings often end < 1 s after the stop (see [[imu-event-labeling]]).
IMU_STOP_ONSET_MIN_SEC = float(os.environ.get('SENSECV_IMU_STOP_ONSET_MIN_SEC', '0.3'))
# A stop cut only fires when the standstill is *confirmed*: it runs to the end
# of the clip (the person does not resume walking) and lasts at least this long,
# so the person verifiably stays put instead of just pausing.
IMU_STOP_CONFIRM_SEC = float(os.environ.get('SENSECV_IMU_STOP_CONFIRM_SEC', '0.8'))
# A slow-down/stop only ever occurs in the clip's final stretch (the person
# halts at the end). Its onset must fall within the last IMU_STOP_TAIL_SEC
# seconds, otherwise it is a mid-clip pause, not the closing stop.
IMU_STOP_TAIL_SEC = float(os.environ.get('SENSECV_IMU_STOP_TAIL_SEC', '3.0'))
# A lateral deviation never occurs at the very start or end of a clip; desvio
# events whose onset/return falls within this margin of either edge are
# rejected as boundary artefacts (camera settling, the closing stop/turn).
IMU_DEVIATION_EDGE_MARGIN_SEC = float(os.environ.get('SENSECV_IMU_DEVIATION_EDGE_MARGIN_SEC', '1.0'))
# Step-cadence detection: a stop is when the step rhythm disappears at the clip
# end. Gait periodicity is measured by autocorrelation over a sliding window;
# a window counts as "stepping" when its peak periodicity reaches this strength.
IMU_STEP_WINDOW_SEC = float(os.environ.get('SENSECV_IMU_STEP_WINDOW_SEC', '2.0'))
IMU_STEP_HOP_SEC = float(os.environ.get('SENSECV_IMU_STEP_HOP_SEC', '0.25'))
IMU_STEP_PERIODICITY = float(os.environ.get('SENSECV_IMU_STEP_PERIODICITY', '0.3'))
# Clips shorter than this can't show a walk-then-stop transition.
IMU_STOP_MIN_CLIP_SEC = float(os.environ.get('SENSECV_IMU_STOP_MIN_CLIP_SEC', '3.5'))
IMU_EXPANDED_MARGIN_SEC = float(os.environ.get('SENSECV_IMU_EXPANDED_MARGIN_SEC', '0.3'))
IMU_DEVIATION_YAW_DEG_S = float(os.environ.get('SENSECV_IMU_DEVIATION_YAW_DEG_S', '25.0'))
IMU_DEVIATION_LAT_VEL = float(os.environ.get('SENSECV_IMU_DEVIATION_LAT_VEL', '0.035'))
IMU_DEVIATION_MERGE_GAP_SEC = float(os.environ.get('SENSECV_IMU_DEVIATION_MERGE_GAP_SEC', '1.5'))
IMU_WALK_STD = 0.5    # m/s², same gait threshold as _orientation_walking_masks
IMU_STOP_STD = 0.25   # m/s², below this the user is considered stopped
IMU_DECEL_RATIO = 0.6  # energy below this fraction of the recent walking baseline


def _mask_runs(mask):
    """Yield (i0, i1) inclusive index runs where mask is True."""
    runs = []
    start = None
    for i, value in enumerate(mask):
        if value and start is None:
            start = i
        elif not value and start is not None:
            runs.append((start, i - 1))
            start = None
    if start is not None:
        runs.append((start, len(mask) - 1))
    return runs


def _stream_rate_stats(times_usec):
    """Sampling-rate health for one sensor stream (PDF section 1)."""
    t = np.asarray(times_usec, dtype=np.float64)
    if len(t) < 3:
        return {'samples': len(t), 'median_hz': 0.0, 'min_hz': 0.0,
                'gap_count': 0, 'jitter_ratio': 0.0}
    dt = np.diff(t) / 1e6
    dt = dt[dt > 0]
    if len(dt) == 0:
        return {'samples': len(t), 'median_hz': 0.0, 'min_hz': 0.0,
                'gap_count': 0, 'jitter_ratio': 0.0}
    median_dt = float(np.median(dt))
    return {
        'samples': int(len(t)),
        'median_hz': round(1.0 / median_dt, 2),
        'min_hz': round(1.0 / float(dt.max()), 2),
        'gap_count': int(np.count_nonzero(dt > 3.0 * median_dt)),
        'jitter_ratio': round(float(dt.std()) / median_dt, 4),
    }


def _saturation_fraction(values, floor):
    """Fraction of samples pinned near the absolute max (clipping evidence)."""
    v = np.abs(np.asarray(values, dtype=np.float64))
    peak = float(v.max()) if len(v) else 0.0
    if peak < floor:
        return 0.0
    return float(np.count_nonzero(v >= 0.98 * peak)) / max(1, len(v))


def _dominant_cadence(amag, fps):
    """Step rate (Hz) and autocorrelation strength of the gait periodicity."""
    x = np.asarray(amag, dtype=np.float64)
    if len(x) < int(fps * 3):
        return None, 0.0
    x = x - x.mean()
    ac = np.correlate(x, x, 'full')[len(x) - 1:]
    if ac[0] <= 1e-9:
        return None, 0.0
    ac = ac / ac[0]
    lo = max(1, int(round(fps * 0.3)))
    hi = min(len(ac) - 1, int(round(fps * 1.2)))
    if hi <= lo:
        return None, 0.0
    lag = lo + int(np.argmax(ac[lo:hi]))
    return round(fps / lag, 2), round(float(ac[lag]), 3)


def _imu_event_series(idx):
    """Frame-aligned signals shared by quality report and event detection."""
    folder = clip_path(idx)
    frames = load_json(folder, 'frames.json')['frames']
    acc = load_json(folder, 'accelerations.json')['accelerations']
    rot = load_json(folder, 'rotations.json')['rotations']

    ft = np.array([f['time_usec'] for f in frames], dtype=np.float64)
    at = np.array([a['time_usec'] for a in acc], dtype=np.float64)
    av = np.array([[a['x'], a['y'], a['z']] for a in acc], dtype=np.float64)
    rt = np.array([r['time_usec'] for r in rot], dtype=np.float64)
    rv = np.array([[r['x'], r['y'], r['z']] for r in rot], dtype=np.float64)
    if len(ft) < 2 or len(av) == 0:
        raise ValueError('sensor data is empty or too short')

    secs = (ft - ft[0]) / 1e6
    fps = float(1e6 / np.median(np.diff(ft)))

    a_at_f = np.array(interp_at_times(at, av, ft.tolist()), dtype=np.float64)
    amag = np.linalg.norm(a_at_f, axis=1)

    W = min(len(ft), max(1, int(fps * 1.0)))
    kernel = np.ones(W) / W
    mean_mag = np.convolve(amag, kernel, mode='same')
    var_mag = np.convolve((amag - mean_mag) ** 2, kernel, mode='same')
    std_mag = np.sqrt(np.maximum(var_mag, 0.0))

    # Gravity-aligned body axes, independent of how the phone is held
    # (em pe / deitado): up = smoothed accel direction, forward = camera -z
    # projected to the horizontal plane, left = up x forward.
    g_smooth_w = max(1, int(round(fps * 0.7)))
    g_dir = np.column_stack([
        _centered_rolling_mean(a_at_f[:, i], g_smooth_w) for i in range(3)
    ])
    g_dir = g_dir / np.maximum(np.linalg.norm(g_dir, axis=1)[:, None], 1e-6)
    minus_z = np.zeros_like(g_dir)
    minus_z[:, 2] = -1.0
    fwd = np.cross(g_dir, np.cross(minus_z, g_dir))
    fwd = fwd / np.maximum(np.linalg.norm(fwd, axis=1)[:, None], 1e-6)
    left = np.cross(g_dir, fwd)

    # Yaw rate around gravity (PDF 2.3: "velocidade angular em yaw").
    yaw_deg_s = np.zeros(len(ft))
    if len(rv):
        r_at_f = np.array(interp_at_times(rt, rv, ft.tolist()), dtype=np.float64)
        yaw_rad_s = np.einsum('ij,ij->i', r_at_f, g_dir)
        smooth_w = max(1, int(round(fps * 0.2)))
        yaw_deg_s = np.degrees(np.array(_centered_rolling_mean(yaw_rad_s, smooth_w)))

    # Lateral velocity (PDF 2.3: "aceleracao lateral persistente"), integrated
    # with rolling-mean drift removal. On gimbal-stabilized captures the gyro
    # barely registers the deviation, but the body still translates sideways;
    # validated at 27/32 correct directions on the IFCE manifest clips.
    lin = a_at_f - g_dir * np.linalg.norm(
        np.column_stack([_centered_rolling_mean(a_at_f[:, i], g_smooth_w) for i in range(3)]),
        axis=1)[:, None]
    lat_acc = np.einsum('ij,ij->i', lin, left)
    lat_vel = np.cumsum(lat_acc) / fps
    drift_w = max(1, int(round(fps * 2.5)))
    lat_vel = lat_vel - np.array(_centered_rolling_mean(lat_vel, drift_w))
    lat_vel = np.array(_centered_rolling_mean(lat_vel, max(1, int(round(fps * 0.6)))))

    return {
        'secs': secs, 'fps': fps,
        'acc_times': at, 'acc_values': av,
        'rot_times': rt, 'rot_values': rv,
        'amag': amag, 'std_mag': std_mag,
        'yaw_deg_s': yaw_deg_s, 'lat_vel': lat_vel,
        'duration_s': float(secs[-1]),
    }


def imu_quality_report(idx):
    """PDF section 1: capture validation checklist with a verdict."""
    series = _imu_event_series(idx)
    acc_rate = _stream_rate_stats(series['acc_times'])
    rot_rate = _stream_rate_stats(series['rot_times'])

    acc_sat = max(_saturation_fraction(series['acc_values'][:, i], 35.0) for i in range(3))
    rot_sat = (max(_saturation_fraction(series['rot_values'][:, i], 15.0) for i in range(3))
               if len(series['rot_values']) else 0.0)

    std_mag = series['std_mag']
    walking = std_mag > IMU_WALK_STD
    stopped = std_mag < IMU_STOP_STD
    walking_frac = float(np.mean(walking))

    cadence_hz, periodicity = None, 0.0
    runs = _mask_runs(walking.tolist())
    if runs:
        i0, i1 = max(runs, key=lambda r: r[1] - r[0])
        cadence_hz, periodicity = _dominant_cadence(series['amag'][i0:i1 + 1], series['fps'])
    if cadence_hz is None:
        # Short clips rarely have a single >=3 s walking run; the clip-level
        # autocorrelation still recovers the step periodicity.
        cadence_hz, periodicity = _dominant_cadence(series['amag'], series['fps'])

    p10, p90 = (np.percentile(std_mag, 10), np.percentile(std_mag, 90)) if len(std_mag) else (0, 0)
    energy_range = float(p90 / max(p10, 1e-6))

    checks = {
        'taxa_amostragem_100hz': acc_rate['median_hz'] >= 100.0,
        'sem_lacunas': acc_rate['gap_count'] == 0,
        'jitter_baixo': acc_rate['jitter_ratio'] < 0.5,
        'sem_saturacao': acc_sat < 0.01 and rot_sat < 0.01,
        'marcha_periodica': periodicity >= 0.3 and cadence_hz is not None,
        'tem_caminhada': walking_frac > 0.2,
        'estados_distinguiveis': energy_range > 2.0 or float(np.mean(stopped)) > 0.05,
    }
    failed = [name for name, ok in checks.items() if not ok]
    critical = {'sem_saturacao', 'tem_caminhada', 'marcha_periodica'}
    if any(name in critical for name in failed):
        verdict = 'rejeitar'
    elif failed:
        verdict = 'baixa_confianca'
    else:
        verdict = 'aceitar'

    return {
        'verdict': verdict,
        'failed_checks': failed,
        'checks': checks,
        'acc_rate': acc_rate,
        'rot_rate': rot_rate,
        'acc_saturation_frac': round(acc_sat, 4),
        'rot_saturation_frac': round(rot_sat, 4),
        'walking_fraction': round(walking_frac, 3),
        'cadence_hz': cadence_hz,
        'gait_periodicity': periodicity,
        'energy_dynamic_range': round(energy_range, 2),
        'duration_s': round(series['duration_s'], 3),
    }


def _event_windows(t1, t2, delta, duration):
    clamp = lambda v: round(min(max(0.0, v), duration), 3)
    return {
        'acao': [clamp(t1), clamp(t2)],
        'decisao': [clamp(t1 - delta), clamp(t1)],
        'expandido': [clamp(t1 - delta), clamp(t2 + IMU_EXPANDED_MARGIN_SEC)],
    }


def detect_imu_events(idx, delta=None):
    """PDF sections 2-5: T1/T2 events with label windows and confidence."""
    delta = IMU_EVENT_DELTA_SEC if delta is None else float(delta)
    series = _imu_event_series(idx)
    quality = imu_quality_report(idx)
    secs, fps = series['secs'], series['fps']
    duration = series['duration_s']
    std_mag, yaw_deg_s = series['std_mag'], series['yaw_deg_s']
    persist_n = max(1, int(round(fps * IMU_EVENT_PERSIST_SEC)))

    events = []

    # Desvio lateral (2.3): a persistent excursion marks the turn away; the
    # deviation is "turn away, then turn back", so consecutive excursions of
    # opposite sign within a short gap merge into one event whose direction is
    # the first excursion and whose T2 is the return to the trajectory.
    # Signal choice: yaw rate when the capture actually rotates (handheld);
    # lateral velocity otherwise (gimbal-stabilized captures barely yaw).
    yaw_strength = float(np.max(np.abs(yaw_deg_s))) / IMU_DEVIATION_YAW_DEG_S if len(yaw_deg_s) else 0.0
    if yaw_strength >= 1.0:
        dev_signal, dev_threshold, dev_source = yaw_deg_s, IMU_DEVIATION_YAW_DEG_S, 'yaw'
    else:
        dev_signal, dev_threshold, dev_source = series['lat_vel'], IMU_DEVIATION_LAT_VEL, 'lateral_velocity'

    turns = []
    # Persistence is tested at 60% of the threshold (the excursion plateau),
    # but the peak must reach the full threshold: rejects single spikes (2.3)
    # without demanding the peak itself lasts the whole persistence window.
    dev_mask = np.abs(dev_signal) > 0.6 * dev_threshold
    for i0, i1 in _mask_runs(dev_mask.tolist()):
        if i1 - i0 + 1 < persist_n:
            continue  # single isolated peaks are explicitly not T1
        segment = dev_signal[i0:i1 + 1]
        if float(np.max(np.abs(segment))) < dev_threshold:
            continue
        turns.append({
            'i0': i0, 'i1': i1,
            'sign': 1.0 if float(np.mean(segment)) > 0 else -1.0,
            'peak': float(np.max(np.abs(segment))),
        })
    merged = []
    for turn in turns:
        prev = merged[-1] if merged else None
        if (prev is not None
                and turn['sign'] != prev['sign']
                and secs[turn['i0']] - secs[prev['i1']] <= IMU_DEVIATION_MERGE_GAP_SEC
                and not prev.get('returned')):
            if turn['peak'] > prev['peak']:
                prev['peak'] = turn['peak']
                prev['peak_sign'] = turn['sign']
            prev['i1'] = turn['i1']
            prev['returned'] = True
            continue
        item = dict(turn)
        item['peak_sign'] = item['sign']
        merged.append(item)
    for turn in merged:
        # Positive = leftward for both signals (right-hand rule around the
        # gravity-aligned up vector; left = up x forward). The strongest
        # excursion of the out-and-back pair carries the deviation direction
        # (validated against the IFCE manifest).
        direction = 'esquerda' if turn['peak_sign'] > 0 else 'direita'
        events.append({
            'type': 'desvio', 'direction': direction,
            'source': dev_source,
            't1': float(secs[turn['i0']]), 't2': float(secs[turn['i1']]),
            'peak': round(turn['peak'], 3),
            'strength': turn['peak'] / dev_threshold,
        })

    # Parada (3.2): gait energy at standstill level for >= IMU_STOP_MIN_SEC.
    stop_mask = std_mag < IMU_STOP_STD
    stop_runs = []
    for i0, i1 in _mask_runs(stop_mask.tolist()):
        if secs[i1] - secs[i0] < IMU_STOP_MIN_SEC:
            continue
        stop_runs.append((i0, i1))
        events.append({
            'type': 'parada', 'direction': None,
            't1': float(secs[i0]), 't2': float(secs[i1]),
            'peak': round(float(np.min(std_mag[i0:i1 + 1])), 3),
            'strength': IMU_STOP_STD / max(float(np.min(std_mag[i0:i1 + 1])), 1e-6),
        })

    # Reducao (3.1): walking energy falls persistently below a fraction of the
    # recent walking baseline without reaching standstill.
    baseline = np.array(_centered_rolling_mean(std_mag, max(1, int(round(fps * 3.0)))))
    decel_mask = (std_mag > IMU_STOP_STD) & (std_mag < IMU_DECEL_RATIO * baseline)
    for i0, i1 in _mask_runs(decel_mask.tolist()):
        if i1 - i0 + 1 < persist_n:
            continue
        in_stop = any(s0 <= i0 <= s1 for s0, s1 in stop_runs)
        if in_stop:
            continue
        ratio = float(np.min(std_mag[i0:i1 + 1] / np.maximum(baseline[i0:i1 + 1], 1e-6)))
        events.append({
            'type': 'reducao', 'direction': None,
            't1': float(secs[i0]), 't2': float(secs[i1]),
            'peak': round(ratio, 3),
            'strength': IMU_DECEL_RATIO / max(ratio, 1e-6),
        })

    events.sort(key=lambda e: e['t1'])

    # Confidence (PDF 1.2 / 5 / 5.1): ambiguity between simultaneous events,
    # weak signal margins, and the capture verdict all lower confidence.
    for i, event in enumerate(events):
        reasons = []
        overlapping = [
            other for j, other in enumerate(events) if j != i
            and other['type'] != event['type']
            and not (other['t2'] < event['t1'] - 0.5 or other['t1'] > event['t2'] + 0.5)
        ]
        # A decel run immediately before a stop is a natural pair, not noise.
        natural_pair = (
            event['type'] in ('reducao', 'parada')
            and all(o['type'] in ('reducao', 'parada') for o in overlapping)
        )
        if quality['verdict'] == 'rejeitar':
            confidence = 'descartar'
            reasons.append('captura IMU rejeitada')
        elif overlapping and not natural_pair:
            confidence = 'baixa'
            reasons.append('eventos sobrepostos/ambiguos')
        elif quality['verdict'] == 'baixa_confianca' or event['strength'] < 1.5:
            confidence = 'baixa'
            if quality['verdict'] == 'baixa_confianca':
                reasons.append('qualidade de captura limitada')
            if event['strength'] < 1.5:
                reasons.append('margem de sinal fraca')
        else:
            confidence = 'alta'
        event['confidence'] = confidence
        event['confidence_reasons'] = reasons
        event['windows'] = _event_windows(event['t1'], event['t2'], delta, duration)
        event['t1'] = round(event['t1'], 3)
        event['t2'] = round(event['t2'], 3)
        event['strength'] = round(event['strength'], 2)

    return {'delta': delta, 'quality': quality, 'events': events,
            'duration_s': round(duration, 3)}


# PDF (criterios_imu_video_orientando) §2.2: the deviation cut is one of three
# label windows around the IMU-detected event (T1 = physical start, T2 = return
# to a stabilized trajectory). The default is the "decisão visual" window
# (T1-Δ → T1) — the scene just before the body reacts, which §2.1/§10 call the
# most adequate target for training a predictive CNN.
DEVIATION_CUT_WINDOW = os.environ.get('SENSECV_DEVIATION_CUT_WINDOW', 'decisao')


def _step_rhythm(amag, fps):
    """Per-window step-cadence strength along the clip.

    Walking is periodic (one accel pulse per step); standing still is not. For
    each sliding window (IMU_STEP_WINDOW_SEC, hop IMU_STEP_HOP_SEC) the
    normalized autocorrelation peak in the step band (period 0.3-1.0 s, i.e.
    1-3.3 Hz cadence) measures how rhythmic that stretch is. Returns
    (centers_s, periodicity) arrays; periodicity >= IMU_STEP_PERIODICITY marks a
    window where the person is taking steps. This ignores the absolute energy,
    so the start/end phone-handling spikes (broadband, non-periodic) do not read
    as walking.
    """
    x = np.asarray(amag, dtype=np.float64)
    n = len(x)
    w = max(8, int(round(fps * IMU_STEP_WINDOW_SEC)))
    hop = max(1, int(round(fps * IMU_STEP_HOP_SEC)))
    lo = max(1, int(round(fps * 0.3)))
    hi = int(round(fps * 1.0))
    centers, per = [], []
    for s in range(0, max(1, n - w + 1), hop):
        seg = x[s:s + w]
        seg = seg - seg.mean()
        ac = np.correlate(seg, seg, 'full')[len(seg) - 1:]
        if ac[0] <= 1e-9:
            p = 0.0
        else:
            ac = ac / ac[0]
            h = min(len(ac) - 1, hi)
            p = float(np.max(ac[lo:h])) if h > lo else 0.0
        centers.append((s + w / 2.0) / fps)
        per.append(p)
    return np.array(centers), np.array(per)


def detect_stop_onset(idx):
    """Time the step rhythm disappears at the clip end, for clips with no desvio.

    A stop is when the person *stops taking steps* before the recording ends:
    the gait cadence is present earlier and then ceases and stays gone through
    the final sample. T1 (`stop_time`) is the onset of that cadence loss — the
    instant stepping stops. It must fall within the clip's final
    IMU_STOP_TAIL_SEC seconds (the stop always closes the clip). Returns None
    when the person is still stepping at the end (a free walk) or never has a
    clear gait.
    """
    series = _imu_event_series(idx)
    fps, amag = series['fps'], series['amag']
    duration = series['duration_s']
    if duration < IMU_STOP_MIN_CLIP_SEC:
        return None  # too short to show a walk-then-stop transition
    centers, per = _step_rhythm(amag, fps)
    if len(per) < 3:
        return None
    stepping = per >= IMU_STEP_PERIODICITY
    if not stepping.any():
        return None  # never a clear gait -> not a "stop"
    if stepping[-1]:
        return None  # still stepping at the end -> free walk, not a stop
    stepping_idx = np.nonzero(stepping)[0]
    onset = float(centers[int(stepping_idx[-1]) + 1])
    if onset < duration - IMU_STOP_TAIL_SEC:
        return None  # cadence loss is not in the clip's closing stretch
    return {'t1': onset, 't2': duration, 'fps': fps, 'duration_s': duration}


def detect_free_walk_span(idx):
    """Whole video for clips with no desvio and no stop — a free walk.

    A clip with a clear gait that never deviates and never stops (the person is
    still stepping at the end, see `detect_stop_onset`) is a *free walk*
    (caminhada livre): unobstructed forward walking start to finish, its own
    dataset class. The cut is the entire video [0, duration]. Returns None only
    when the clip has no discernible gait at all.
    """
    series = _imu_event_series(idx)
    fps, amag = series['fps'], series['amag']
    _, per = _step_rhythm(amag, fps)
    if len(per) == 0 or not (per >= IMU_STEP_PERIODICITY).any():
        return None  # no gait at all
    return {
        't1': 0.0, 't2': float(series['duration_s']),
        'fps': fps, 'duration_s': float(series['duration_s']),
    }


# ─── Collection manifest labels (data/labels/*.xlsx) ─────────────────────────
# Each collection ships an .xlsx where every row is one recorded clip. The clip's
# ground-truth class lives in fixed columns (consistent across every sheet):
#   col 0  ID                  -> clip number (matches the clip folder 01,02,…)
#   col 2  POSIÇÃO OBSTACULO   -> obstacle position, or "PARADA" / "LIVRE"
#   col 3  DESVIO              -> "PARA DIREITA" / "PARA ESQUERDA" / "SEM DESVIO"
# We key these by recording date (parsed from the filename) + clip number so a
# clip's display name (SenseCV-02-06-2026-…/…/07) maps straight to its row.
LABELS_DIR = os.path.join(DATA_DIR, 'labels')
_MANIFEST_LABELS_CACHE = None
_PT_MONTHS = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'abril': 4, 'maio': 5, 'junho': 6,
    'julho': 7, 'agosto': 8, 'setembro': 9, 'outubro': 10, 'novembro': 11,
    'dezembro': 12,
}


def _ascii_upper(value):
    """Strip accents (the sheets are full of mojibake) and upper-case."""
    import unicodedata
    if value is None:
        return ''
    norm = unicodedata.normalize('NFKD', str(value))
    return ''.join(c for c in norm if not unicodedata.combining(c)).upper().strip()


def _manifest_date_key(text):
    """Extract a DD-MM-YYYY key from a label filename ('… 02 de junho de 2026 …')."""
    m = re.search(r'(\d{1,2})\s+de\s+([a-zç]+)\s+de\s+(\d{4})',
                  _ascii_upper(text).lower())
    if not m:
        return None
    month = _PT_MONTHS.get(m.group(2))
    if not month:
        return None
    return f'{int(m.group(1)):02d}-{month:02d}-{m.group(3)}'


def _manifest_row_kind(desvio_cell, pos_cell):
    """Map a manifest row's DESVIO + POSIÇÃO cells to (event_type, side)."""
    desvio = _ascii_upper(desvio_cell)
    pos = _ascii_upper(pos_cell)
    if 'PARA DIREITA' in desvio:
        return ('desvio', 'RIGHT')
    if 'PARA ESQUERDA' in desvio:
        return ('desvio', 'LEFT')
    if 'LIVRE' in pos:
        return ('livre', 'NONE')
    if 'PARADA' in pos:
        return ('parada', 'NONE')
    # "SEM DESVIO" with the obstacle centered = walk up to it and stop.
    if 'SEM DESVIO' in desvio:
        return ('parada', 'NONE')
    return None


def _load_manifest_labels():
    """Parse every data/labels/*.xlsx into {date_key: {clip_id: (event_type, side)}}."""
    global _MANIFEST_LABELS_CACHE
    if _MANIFEST_LABELS_CACHE is not None:
        return _MANIFEST_LABELS_CACHE
    out = {}
    try:
        import openpyxl
    except Exception:
        _MANIFEST_LABELS_CACHE = out
        return out
    if os.path.isdir(LABELS_DIR):
        for fn in sorted(os.listdir(LABELS_DIR)):
            if not fn.lower().endswith('.xlsx') or fn.startswith('~$'):
                continue
            date_key = _manifest_date_key(fn)
            if not date_key:
                continue
            try:
                wb = openpyxl.load_workbook(
                    os.path.join(LABELS_DIR, fn), read_only=True, data_only=True)
            except Exception:
                continue
            clips = out.setdefault(date_key, {})
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    try:
                        clip_id = int(float(row[0]))  # skips header / info rows
                    except (TypeError, ValueError):
                        continue
                    kind = _manifest_row_kind(
                        row[3] if len(row) > 3 else None,
                        row[2] if len(row) > 2 else None)
                    if kind:
                        clips.setdefault(clip_id, kind)
            wb.close()
    _MANIFEST_LABELS_CACHE = out
    return out


def _manifest_kind_for_clip(idx):
    """Ground-truth (event_type, side) for a clip from its collection manifest.

    Resolves the recording date and clip number from the display name
    (SenseCV-02-06-2026-IFCE/…/07 -> date 02-06-2026, clip 7). Returns None when
    there is no manifest row for the clip.
    """
    try:
        display = CLIPS[idx]
    except (IndexError, TypeError):
        return None
    parts = str(display).split('/')
    dm = re.search(r'(\d{2})-(\d{2})-(\d{4})', parts[0])
    cm = re.search(r'(\d+)', parts[-1])
    if not dm or not cm:
        return None
    date_key = f'{dm.group(1)}-{dm.group(2)}-{dm.group(3)}'
    return _load_manifest_labels().get(date_key, {}).get(int(cm.group(1)))


def _desvio_window_from_signal(idx, side, delta):
    """Locate a deviation's [T1, T2] directly from the lateral-velocity signal.

    Used for manifest-confirmed desvios that the generic threshold detector
    misses — typically the BECE/UFC collections (days 20-22), whose lateral
    excursions are real but smaller in absolute terms than the IFCE clips the
    fixed threshold was tuned on. The manifest already gives the side, so we look
    for the strongest sustained excursion in that known direction (no absolute
    threshold) and frame the out-and-back around it.

    Returns (t1, t2, direction, strength) or None.
    """
    try:
        series = _imu_event_series(idx)
    except Exception:
        return None
    lat = np.asarray(series['lat_vel'], dtype=np.float64)
    secs = np.asarray(series['secs'], dtype=np.float64)
    if len(lat) < 3:
        return None
    # Positive lat_vel = leftward (same convention as detect_imu_events). Orient
    # the signal so a deviation in the manifest's direction reads positive.
    sign = 1.0 if side == 'LEFT' else -1.0
    sig = lat * sign
    peak_i = int(np.argmax(sig))
    if sig[peak_i] <= 0:
        # No excursion in the expected direction (sensor sign/orientation odd):
        # fall back to the largest lateral excursion either way — the manifest
        # still owns the side, we only need the timing.
        peak_i = int(np.argmax(np.abs(lat)))
        peak_val = float(abs(lat[peak_i]))
        prof = np.abs(lat)
    else:
        peak_val = float(sig[peak_i])
        prof = sig
    if peak_val <= 1e-9:
        return None
    # Onset/return: the contiguous run around the peak where the excursion stays
    # above 30% of its own height (relative, so it adapts to each collection).
    thr = 0.30 * peak_val
    i0 = peak_i
    while i0 > 0 and prof[i0 - 1] > thr:
        i0 -= 1
    i1 = peak_i
    while i1 < len(prof) - 1 and prof[i1 + 1] > thr:
        i1 += 1
    t1, t2 = float(secs[i0]), float(secs[i1])
    direction = 'esquerda' if side == 'LEFT' else 'direita'
    strength = peak_val / IMU_DEVIATION_LAT_VEL
    return t1, t2, direction, strength


def suggest_deviation_cut(idx, delta=None, window=None):
    """Suggested deviation cut from the IMU event windows (PDF §2.2/§2.3).

    Picks the strongest `desvio` event from `detect_imu_events` and returns the
    requested label window:
      'decisao'   -> [T1-Δ, T1]          scene before the reaction (default)
      'acao'      -> [T1, T2]            the deviation execution
      'expandido' -> [T1-Δ, T2+margem]   both, robust to sync uncertainty
    Side (LEFT/RIGHT) comes from the same event's direction, so cut and label
    are always consistent.

    Calibration (per recording convention):
      - desvio never sits at the very start or end of the clip; edge events
        within IMU_DEVIATION_EDGE_MARGIN_SEC of either boundary are dropped.
      - parada only ever closes the clip (onset in the last IMU_STOP_TAIL_SEC s,
        enforced in `detect_stop_onset`).
      - livre is the only cut that spans the *entire* video [0, duration].

    Two fallbacks when there is no desvio: first the stop onset (`parada`,
    `detect_stop_onset`); then, with neither desvio nor parada, the clip is a
    free walk (`livre`, `detect_free_walk_span`) and the whole video is
    returned. Only a clip with no usable walking yields found=False.

    When the clip has a row in the collection manifest (data/labels/*.xlsx), that
    ground-truth class wins: it picks the cut branch (desvio/parada/livre) and the
    deviation side, instead of the IMU's guess. The IMU is still used for the cut
    *timing* within the chosen class.
    """
    window = window or DEVIATION_CUT_WINDOW
    try:
        result = detect_imu_events(idx, delta=delta)
    except Exception as e:
        return {'found': False, 'mode': 'deviation', 'has_vertical': True,
                'message': f'IMU indisponivel: {e}'}
    duration_clip = float(result.get('duration_s') or 0.0)
    margin = IMU_DEVIATION_EDGE_MARGIN_SEC

    mk = _manifest_kind_for_clip(idx)
    mk_type, mk_side = (mk if mk else (None, None))

    all_desvios = [e for e in result.get('events', []) if e.get('type') == 'desvio']
    # A deviation never happens at the very start or end of the clip; drop any
    # desvio whose onset/return sits inside the edge margin (only when the clip
    # is long enough that a margin on both sides still leaves a middle).
    desvios = all_desvios
    if duration_clip > 2 * margin:
        desvios = [
            e for e in all_desvios
            if float(e.get('t1', 0.0)) >= margin
            and float(e.get('t2', duration_clip)) <= duration_clip - margin
        ]

    def _desvio_cut(side_override=None, relax=False):
        # `relax` lets a manifest-confirmed desvio fall back to edge events the
        # default path would have dropped, so the known class still yields a cut.
        pool = desvios or (all_desvios if relax else [])
        if not pool:
            # Manifest confirms a desvio but the threshold detector found no event
            # (smaller excursions on the BECE/UFC days). Recover the timing from
            # the lateral-velocity signal in the manifest's known direction.
            if side_override and duration_clip > 0:
                delta_v = result.get('delta') or IMU_EVENT_DELTA_SEC
                sig = _desvio_window_from_signal(idx, side_override, delta_v)
                if sig is not None:
                    t1, t2, direction, strength = sig
                    windows = _event_windows(t1, t2, delta_v, duration_clip)
                    win = windows.get(window) or windows.get('decisao')
                    start, end = float(win[0]), float(win[1])
                    msg = 'desvio confirmado pelo manifesto; janela do sinal lateral'
                else:
                    # No usable signal at all: frame the decision window around the
                    # clip midpoint (deviation happens mid-clip by convention).
                    mid = duration_clip / 2.0
                    start = max(0.0, mid - delta_v / 2.0)
                    end = min(duration_clip, start + delta_v)
                    t1, t2 = start, end
                    direction = {'LEFT': 'esquerda', 'RIGHT': 'direita'}.get(side_override)
                    strength = None
                    msg = 'desvio confirmado pelo manifesto; janela aproximada'
                if end - start < 0.2:
                    start = max(0.0, end - 0.2)
                return {
                    'found': True, 'mode': 'deviation', 'window': window,
                    'has_vertical': True, 'event_type': 'desvio',
                    'start': round(start, 3), 'end': round(end, 3),
                    't1': round(t1, 3), 't2': round(t2, 3),
                    'delta': delta_v, 'direction': direction, 'side': side_override,
                    'strength': strength, 'label_source': 'manifesto', 'message': msg,
                }
            return None
        best = max(pool, key=lambda e: e.get('strength', 0.0))
        windows = best.get('windows', {})
        win = windows.get(window) or windows.get('decisao') or [best['t1'], best['t2']]
        start, end = float(win[0]), float(win[1])
        if end - start < 0.2:  # near clip-start truncation: keep a usable minimum
            start = max(0.0, end - 0.2)
        side = side_override or {'esquerda': 'LEFT', 'direita': 'RIGHT'}.get(
            best.get('direction'), 'NONE')
        return {
            'found': True, 'mode': 'deviation', 'window': window, 'has_vertical': True,
            'event_type': 'desvio',
            'start': round(start, 3), 'end': round(end, 3),
            't1': best['t1'], 't2': best['t2'], 'delta': result.get('delta'),
            'direction': best.get('direction'), 'side': side,
            'confidence': best.get('confidence'), 'strength': best.get('strength'),
            'label_source': 'manifesto' if side_override else 'imu',
        }

    def _parada_cut(force=False):
        # Cut the same PDF label window around the stop onset (T1), so a
        # 'decisao' cut frames the scene just before the halt.
        try:
            stop = detect_stop_onset(idx)
        except Exception:
            stop = None
        if stop is None:
            # Manifest confirms a parada but the IMU stop detector was unsure: by
            # the recording convention the halt closes the clip, so frame the
            # decision window at the very end.
            if force and duration_clip > 0:
                delta_v = result.get('delta') or IMU_EVENT_DELTA_SEC
                end = duration_clip
                start = max(0.0, end - delta_v)
                return {
                    'found': True, 'mode': 'deviation', 'window': window,
                    'has_vertical': True, 'event_type': 'parada',
                    'start': round(start, 3), 'end': round(end, 3),
                    't1': round(end, 3), 't2': round(end, 3),
                    'delta': delta_v, 'direction': None, 'side': 'NONE',
                    'stop_time': round(end, 3),
                    'message': 'parada confirmada pelo manifesto; corte no fim do clipe',
                }
            return None
        delta_v = result.get('delta') or IMU_EVENT_DELTA_SEC
        windows = _event_windows(stop['t1'], stop['t2'], delta_v, stop['duration_s'])
        win = windows.get(window) or windows.get('decisao')
        start, end = float(win[0]), float(win[1])
        if end - start < 0.2:  # near clip-start truncation: usable minimum
            start = max(0.0, end - 0.2)
        return {
            'found': True, 'mode': 'deviation', 'window': window,
            'has_vertical': True, 'event_type': 'parada',
            'start': round(start, 3), 'end': round(end, 3),
            't1': round(stop['t1'], 3), 't2': round(stop['t2'], 3),
            'delta': delta_v, 'direction': None, 'side': 'NONE',
            'stop_time': round(stop['t1'], 3),
            'message': 'corte na parada (inicio da parada)',
        }

    def _livre_cut(force=False):
        # Free walk (caminhada livre): the only class whose cut spans the *entire*
        # video. `force` keeps the whole clip even when the gait detector is
        # unsure, used when the manifest already confirms it is a free walk.
        try:
            free = detect_free_walk_span(idx)
        except Exception:
            free = None
        if free is None and not force:
            return None
        dur = float(free['duration_s']) if free is not None else duration_clip
        if dur <= 0:
            return None
        start, end = 0.0, dur
        if end - start < 0.2:  # degenerate clip: keep a usable minimum
            end = start + 0.2
        return {
            'found': True, 'mode': 'deviation', 'window': window,
            'has_vertical': True, 'event_type': 'livre',
            'start': round(start, 3), 'end': round(end, 3),
            't1': round(start, 3), 't2': round(end, 3),
            'delta': result.get('delta'), 'direction': None, 'side': 'NONE',
            'message': 'caminhada livre (trecho util)',
        }

    not_found = {'found': False, 'mode': 'deviation', 'has_vertical': True,
                 'message': 'nenhum desvio, parada ou caminhada detectados pela IMU'}

    # Manifest ground truth decides the cut branch when the clip has a row.
    if mk_type == 'livre':
        return _livre_cut(force=True) or _parada_cut() or _desvio_cut(relax=True) or not_found
    if mk_type == 'parada':
        return _parada_cut(force=True) or _livre_cut(force=True) or not_found
    if mk_type == 'desvio':
        return (_desvio_cut(side_override=mk_side, relax=True)
                or _parada_cut() or _livre_cut(force=True) or not_found)

    # No manifest entry: keep the original IMU-first order (desvio→parada→livre).
    return _desvio_cut() or _parada_cut() or _livre_cut() or not_found


def _safe_clip_name(display):
    return display.replace('/', '_').replace('\\', '_')


def _ffmpeg_cut(src_video, start, end, out_video):
    import imageio_ffmpeg
    ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    cmd = [
        ffmpeg, '-y',
        '-ss', f'{start:.3f}',
        '-noautorotate',
        '-i', src_video,
        '-t', f'{end-start:.3f}',
        '-map', '0:v:0',
        '-an',
        '-map_metadata', '0',
        '-c:v', 'libx264',
        '-preset', 'ultrafast',
        '-crf', '23',
        '-pix_fmt', 'yuv420p',
        '-movflags', '+faststart',
        out_video,
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(r.stderr[-500:])


def _derived_file_url(path):
    rel = os.path.relpath(os.path.abspath(path), os.path.abspath(DERIVED_DIR))
    return '/derived-file/' + quote(rel.replace(os.sep, '/'))


def _write_export_review_video(video_paths, out_video, fps=SSIM_REVIEW_FPS):
    """Write one MP4 containing every frame from the exported videos."""
    try:
        import cv2
    except Exception as e:
        raise RuntimeError(f'OpenCV indisponivel para video de revisao: {e}')
    try:
        import imageio_ffmpeg
        ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    except Exception as e:
        raise RuntimeError(f'ffmpeg indisponivel para video de revisao: {e}')

    video_paths = [p for p in video_paths if p and os.path.isfile(p)]
    if not video_paths:
        return 0

    first = cv2.VideoCapture(video_paths[0])
    if not first.isOpened():
        raise RuntimeError('primeiro video exportado nao pode ser lido')
    width, height = _review_frame_size(
        first.get(cv2.CAP_PROP_FRAME_WIDTH),
        first.get(cv2.CAP_PROP_FRAME_HEIGHT),
    )
    first.release()

    os.makedirs(os.path.dirname(out_video), exist_ok=True)
    cmd = [
        ffmpeg, '-y',
        '-f', 'rawvideo',
        '-vcodec', 'rawvideo',
        '-pix_fmt', 'rgb24',
        '-s', f'{width}x{height}',
        '-r', f'{fps}',
        '-i', '-',
        '-an',
        '-c:v', 'libx264',
        '-preset', 'ultrafast',
        '-crf', '23',
        '-pix_fmt', 'yuv420p',
        '-movflags', '+faststart',
        out_video,
    ]
    proc = subprocess.Popen(cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    written = 0
    try:
        for path in video_paths:
            cap = cv2.VideoCapture(path)
            if not cap.isOpened():
                continue
            while True:
                ok, frame = cap.read()
                if not ok:
                    break
                if frame.shape[1] != width or frame.shape[0] != height:
                    frame = cv2.resize(frame, (width, height), interpolation=cv2.INTER_AREA)
                proc.stdin.write(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB).tobytes())
                written += 1
            cap.release()
    finally:
        if proc.stdin:
            proc.stdin.close()
    stderr = proc.stderr.read().decode('utf-8', errors='replace')
    proc.wait()
    if proc.returncode != 0:
        raise RuntimeError(stderr[-500:])
    return written


def export_set(indices, out_dir, mode='walking', verbose=False, build_review=True):
    """Batch-export an auto-suggested cut per clip into `out_dir`.

    mode='walking' uses suggest_crop(idx, 'walking'); mode='lateral' uses
    suggest_lateral_deviation(idx). Writes a sources.csv ledger and returns a
    summary dict. Idempotent: per-clip subfolders are wiped before re-cutting.
    """
    import csv
    os.makedirs(out_dir, exist_ok=True)
    csv_path = os.path.join(out_dir, 'sources.csv')
    rows = []
    exported_videos = []
    n_ok = n_skip = n_fail = 0

    for idx in indices:
        display    = CLIPS[idx]
        src_folder = CLIP_PATHS[display]
        if '/' in display:
            root_label, sub = display.split('/', 1)
        else:
            root_label, sub = 'supermercado', display
        name = _safe_clip_name(display)
        out_folder = os.path.join(out_dir, name)
        out_video  = os.path.join(out_folder, name + '.mp4')

        row = {'output_folder': name, 'source_display': display,
               'source_root': root_label, 'source_subfolder': sub,
               'source_path': src_folder, 'mode': mode,
               'start': '', 'end': '', 'duration': '',
               'frames_before': '', 'frames_after': '',
               'ssim_threshold': '', 'ssim_status': '', 'status': ''}

        if mode == 'deviation':
            sug = suggest_deviation_cut(idx)
        elif mode == 'lateral':
            sug = suggest_lateral_deviation(idx)
        else:
            sug = suggest_crop(idx, 'walking')

        if not sug.get('found'):
            row['status'] = 'no_segment: ' + str(sug.get('message', ''))
            rows.append(row); n_skip += 1
            if verbose: print(f'  [skip] {display}: {sug.get("message","")}')
            continue

        start, end = float(sug['start']), float(sug['end'])
        row.update(start=f'{start:.3f}', end=f'{end:.3f}',
                   duration=f'{end-start:.3f}')
        selection = ssim_frame_selection(idx, start, end)
        row.update(
            frames_before=selection.get('frames_before', ''),
            frames_after=selection.get('frames_after', ''),
            ssim_threshold=selection.get('ssim_threshold', ''),
            ssim_status=selection.get('error', 'ok'),
        )

        if os.path.isdir(out_folder):
            import shutil; shutil.rmtree(out_folder, ignore_errors=True)
        os.makedirs(out_folder)

        try:
            _ffmpeg_cut(clip_video_path(idx), start, end, out_video)
        except Exception as e:
            import shutil; shutil.rmtree(out_folder, ignore_errors=True)
            row['status'] = f'ffmpeg_error: {e}'
            rows.append(row); n_fail += 1
            if verbose: print(f'  [fail] {display}: {e}')
            continue

        sensor_error = None
        try:
            save_sensor_data(idx, start, end, out_folder)
        except Exception as e:
            sensor_error = str(e)

        try:
            save_ssim_selection(selection, out_folder)
            save_ssim_review_videos(idx, start, end, selection, out_folder)
        except Exception as e:
            row['ssim_status'] = f"{row['ssim_status']}; review_error: {e}"

        if sensor_error:
            row['status'] = f'ok_no_sensor: {sensor_error}'
        else:
            row['status'] = 'ok'

        exported_videos.append(out_video)
        rows.append(row); n_ok += 1
        if verbose: print(f'  [ok]   {display}: {start:.2f}-{end:.2f}s')

    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=[
            'output_folder', 'source_display', 'source_root',
            'source_subfolder', 'source_path', 'mode',
            'start', 'end', 'duration', 'frames_before', 'frames_after',
            'ssim_threshold', 'ssim_status', 'status'])
        w.writeheader(); w.writerows(rows)

    review = None
    if exported_videos and build_review:
        review_path = os.path.join(out_dir, 'review_all_frames.mp4')
        try:
            frame_count = _write_export_review_video(exported_videos, review_path)
            review = {
                'path': review_path,
                'url': _derived_file_url(review_path),
                'frames': frame_count,
                'videos': len(exported_videos),
            }
        except Exception as e:
            review = {'error': str(e), 'videos': len(exported_videos)}

    return {'ok': n_ok, 'skipped': n_skip, 'failed': n_fail,
            'total': len(indices), 'out_dir': out_dir, 'csv_path': csv_path,
            'csv_url': _derived_file_url(csv_path), 'review': review}


def _preset_filter(preset):
    if preset == 'all':
        return lambda name: not name.startswith('exports/')
    if preset == 'sensecv':
        return lambda name: name in WALKING_ONLY and not name.startswith('exports/')
    if preset == 'supermarket':
        return lambda name: name not in WALKING_ONLY and not name.startswith('exports/')
    return None


def _preset_out_dir(preset, mode):
    base = {'all': 'Todos os videos', 'sensecv': 'SenseCV', 'supermarket': 'Supermercado'}[preset]
    suffix = '' if mode == 'walking' else ' (lateral)'
    return os.path.join(DERIVED_DIR, f'{base} dataset{suffix}')


PRESETS = ('all', 'sensecv', 'supermarket')


# ─── SenseCV two-head keras model (live inference) ───────────────────────────
# Mirrors the lazy-loaded DroNet runtime: the local two-head .keras model is
# loaded on first use and degrades gracefully (controlled in-payload error)
# when TensorFlow or the model file is unavailable. Heads are matched by output
# width: 2 classes -> obstacle (NONE/OBSTACLE), 3 classes -> deviation
# (LEFT/RIGHT/NONE).
# The default model ships at PROJECT_ROOT/best_model.keras, but the operator can
# upload any other .keras file at runtime via /api/upload-model. The active path
# lives in `_sensemodel_runtime['path']` so it can be swapped without a restart;
# SENSECV_MODEL_PATH remains the env-configurable default / "padrão" fallback.
SENSECV_MODEL_PATH = os.environ.get(
    'SENSECV_MODEL_PATH', os.path.join(PROJECT_ROOT, 'best_model.keras'))
SENSECV_OBSTACLE_LABELS = ['NONE', 'OBSTACLE']
SENSECV_DEVIATION_LABELS = ['LEFT', 'RIGHT', 'NONE']
_sensemodel_runtime = {'model': None, 'error': None, 'path': SENSECV_MODEL_PATH}
_sensemodel_cache = OrderedDict()


def _set_sensemodel_path(path):
    """Point live inference at a new .keras file and drop the cached runtime.

    The next /api/sensemodel request lazily loads the new model. Cached
    per-frame predictions are cleared so they cannot leak across models.
    """
    _sensemodel_runtime['model'] = None
    _sensemodel_runtime['error'] = None
    _sensemodel_runtime['path'] = path
    _sensemodel_cache.clear()


def _load_sensemodel_runtime():
    if _sensemodel_runtime['model'] is not None or _sensemodel_runtime['error'] is not None:
        return _sensemodel_runtime
    try:
        path = _sensemodel_runtime['path']
        if not os.path.isfile(path):
            raise FileNotFoundError(f'modelo .keras nao encontrado: {path}')
        os.environ.setdefault('TF_CPP_MIN_LOG_LEVEL', '3')
        from tensorflow import keras
        _sensemodel_runtime['model'] = keras.models.load_model(path, compile=False)
    except Exception as e:
        _sensemodel_runtime['error'] = str(e)
    return _sensemodel_runtime


def _sensemodel_input_shape(model):
    shape = model.inputs[0].shape
    dims = [int(d) for d in shape[1:] if d is not None]
    if len(dims) == 3:
        h, w, c = dims
    elif len(dims) == 2:
        h, w, c = dims[0], dims[1], 1
    else:
        h, w, c = 224, 224, 3
    return h, w, c


def _label_for(vec):
    vec = list(vec)
    i = int(np.argmax(vec)) if vec else 0
    labels = SENSECV_OBSTACLE_LABELS if len(vec) == 2 else SENSECV_DEVIATION_LABELS
    label = labels[i] if i < len(labels) else str(i)
    prob = float(vec[i]) if vec else 0.0
    return label, prob


def _sensemodel_preprocess(frame_bgr, model):
    """Resize/colour-convert one BGR frame into the model's float32 input array.

    The exported MobileNetV2 has an embedded Rescaling(1/127.5, -1) layer, so we
    feed raw 0-255 pixels here; pre-dividing by 255 collapses the input range.
    """
    import cv2
    h, w, c = _sensemodel_input_shape(model)
    img = cv2.resize(frame_bgr, (w, h), interpolation=cv2.INTER_AREA)
    if c == 1:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)[..., None]
    else:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    return img.astype('float32')


def _split_head_arrays(preds):
    """Map raw model outputs to (obstacle_array, deviation_array) by output width.

    Each returned array is shaped (batch, classes): width 2 -> obstacle
    (NONE/OBSTACLE), width 3 -> deviation (LEFT/RIGHT/NONE).
    """
    if not isinstance(preds, (list, tuple)):
        preds = [preds]
    arrs = [np.atleast_2d(np.asarray(p)) for p in preds]
    obstacle_arr = next((a for a in arrs if a.shape[-1] == 2), arrs[0])
    deviation_arr = next((a for a in arrs if a.shape[-1] == 3), arrs[-1])
    return obstacle_arr, deviation_arr


def sensemodel_frame_classification(idx, time_s, exact=False):
    runtime = _load_sensemodel_runtime()
    if runtime['error']:
        return {'available': False, 'error': runtime['error']}
    model = runtime['model']

    try:
        import cv2
    except Exception as e:
        return {'available': False, 'error': f'OpenCV indisponivel: {e}'}

    video_path = clip_video_path(idx)
    mtime = os.path.getmtime(video_path)
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        return {'available': False, 'error': 'video unreadable'}
    fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    if fps <= 0 or frame_count <= 0:
        cap.release()
        return {'available': False, 'error': 'invalid video metadata'}

    duration = frame_count / fps
    sample_time = max(0.0, min(float(time_s or 0.0), duration))
    if not exact:
        sample_time = math.floor(sample_time * DRONET_SAMPLE_FPS) / DRONET_SAMPLE_FPS
    frame_idx = min(frame_count - 1, max(0, int(round(sample_time * fps))))

    key = (CLIPS[idx], mtime, frame_idx)
    if key in _sensemodel_cache:
        _sensemodel_cache.move_to_end(key)
        cap.release()
        return _sensemodel_cache[key]

    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
    ok, frame = cap.read()
    cap.release()
    if not ok:
        return {'available': False, 'error': 'frame decode failed', 'frame': frame_idx}

    batch = _sensemodel_preprocess(frame, model)[None, ...]
    preds = model.predict(batch, verbose=0)
    obstacle_arr, deviation_arr = _split_head_arrays(preds)
    obstacle_vec = obstacle_arr[0].tolist()
    deviation_vec = deviation_arr[0].tolist()
    obstacle_label, obstacle_prob = _label_for(obstacle_vec)
    deviation_label, deviation_prob = _label_for(deviation_vec)

    result = {
        'available': True,
        'clip': CLIPS[idx],
        'frame': frame_idx,
        'time_s': round(frame_idx / fps, 4),
        'requested_time_s': round(float(time_s or 0.0), 4),
        'sample_fps': DRONET_SAMPLE_FPS,
        'exact': bool(exact),
        'source_fps': round(fps, 4),
        'model': os.path.basename(_sensemodel_runtime['path']),
        'obstacle_label': obstacle_label,
        'obstacle_prob': obstacle_prob,
        'deviation_label': deviation_label,
        'deviation_prob': deviation_prob,
        'obstacle_probs': obstacle_vec,
        'deviation_probs': deviation_vec,
    }
    _sensemodel_cache[key] = result
    if len(_sensemodel_cache) > 64:
        _sensemodel_cache.popitem(last=False)
    return result


# ─── Class-activation (Grad-CAM) maps ────────────────────────────────────────
# Grad-CAM heatmaps for the live frame, exposing where each model "looks":
#   • DroNet collision (classification) head           -> COLLISION/CLEAR
#   • SenseCV two-head .keras model, obstacle head 1   -> NONE/OBSTACLE
#   • SenseCV two-head .keras model, deviation head 2  -> LEFT/RIGHT/NONE
# Each map is returned as a self-contained PNG data URI (heatmap blended over the
# exact pixels that model received), so the viewer can show them side by side.
_activation_cache = OrderedDict()


def _cam_to_data_uri(cam, base_bgr, alpha=0.5):
    """Normalize a (H,W) CAM, colorize it, blend over a BGR base, return a PNG data URI."""
    import cv2
    import base64
    cam = np.asarray(cam, dtype=np.float32)
    cam = cam - float(cam.min())
    peak = float(cam.max())
    if peak > 1e-8:
        cam = cam / peak
    cam_u8 = (cam * 255.0).astype('uint8')
    h, w = base_bgr.shape[:2]
    cam_u8 = cv2.resize(cam_u8, (w, h), interpolation=cv2.INTER_CUBIC)
    heat = cv2.applyColorMap(cam_u8, cv2.COLORMAP_JET)
    overlay = cv2.addWeighted(base_bgr, 1.0 - alpha, heat, alpha, 0.0)
    ok, buf = cv2.imencode('.png', overlay)
    if not ok:
        raise RuntimeError('falha ao codificar PNG do mapa de ativacao')
    return 'data:image/png;base64,' + base64.b64encode(buf.tobytes()).decode('ascii')


def _dronet_activation(frame_bgr):
    """Grad-CAM over DroNet's last conv layer for the collision (classification) head."""
    runtime = _load_dronet_runtime()
    if runtime['error']:
        return {'available': False, 'error': runtime['error']}
    torch = runtime['torch']
    model = runtime['model']
    try:
        import cv2
        tensor, crop = runtime['preprocess_bgr'](frame_bgr)
        target = model.conv9.conv  # last 3x3 conv in residual block 3
        store = {}
        h_fwd = target.register_forward_hook(
            lambda m, i, o: store.__setitem__('act', o))
        h_bwd = target.register_full_backward_hook(
            lambda m, gi, go: store.__setitem__('grad', go[0]))
        try:
            model.zero_grad(set_to_none=True)
            _steer, coll = model(tensor)
            coll.backward()
            act = store['act'][0].detach()    # (C, H, W)
            grad = store['grad'][0].detach()  # (C, H, W)
        finally:
            h_fwd.remove()
            h_bwd.remove()
        weights = grad.mean(dim=(1, 2))
        cam = torch.relu((weights[:, None, None] * act).sum(dim=0)).cpu().numpy()
        prob = float(coll.item())
        base = cv2.cvtColor(crop, cv2.COLOR_GRAY2BGR)
        return {
            'available': True,
            'label': 'COLLISION' if prob >= 0.5 else 'CLEAR',
            'prob': prob,
            'image': _cam_to_data_uri(cam, base),
        }
    except Exception as e:
        return {'available': False, 'error': str(e)}


def _keras_last_conv_layer(model):
    """Last layer with a 4D (B,H,W,C) output — the conv feature map for Grad-CAM."""
    for layer in reversed(model.layers):
        try:
            shape = layer.output.shape
        except Exception:
            continue
        if shape is not None and len(shape) == 4:
            return layer
    return None


def _sensemodel_activation(frame_bgr):
    """Grad-CAM for the two-head .keras model: obstacle (head 1) and deviation (head 2)."""
    runtime = _load_sensemodel_runtime()
    if runtime['error']:
        err = {'available': False, 'error': runtime['error']}
        return err, dict(err)
    model = runtime['model']
    try:
        import cv2
        import tensorflow as tf
        last_conv = _keras_last_conv_layer(model)
        if last_conv is None:
            raise RuntimeError('nenhuma camada convolucional encontrada no modelo')

        inp = _sensemodel_preprocess(frame_bgr, model)[None, ...]  # raw 0-255
        disp = inp[0]
        if disp.shape[-1] == 1:
            base = cv2.cvtColor(disp.astype('uint8'), cv2.COLOR_GRAY2BGR)
        else:
            base = cv2.cvtColor(disp.astype('uint8'), cv2.COLOR_RGB2BGR)

        grad_model = tf.keras.Model(model.inputs,
                                    [last_conv.output] + list(model.outputs))
        x = tf.convert_to_tensor(inp)
        with tf.GradientTape(persistent=True) as tape:
            outputs = grad_model(x, training=False)
            conv_out = outputs[0]
            preds = outputs[1:]
            obstacle_t = next((p for p in preds if p.shape[-1] == 2), preds[0])
            deviation_t = next((p for p in preds if p.shape[-1] == 3), preds[-1])
            obs_idx = int(tf.argmax(obstacle_t[0]))
            dev_idx = int(tf.argmax(deviation_t[0]))
            obs_target = obstacle_t[0, obs_idx]
            dev_target = deviation_t[0, dev_idx]

        def _cam(target):
            grads = tape.gradient(target, conv_out)
            weights = tf.reduce_mean(grads, axis=(0, 1, 2))
            cam = tf.reduce_sum(conv_out[0] * weights, axis=-1)
            return tf.nn.relu(cam).numpy()

        obs_cam = _cam(obs_target)
        dev_cam = _cam(dev_target)
        del tape

        obs_vec = obstacle_t[0].numpy().tolist()
        dev_vec = deviation_t[0].numpy().tolist()
        obstacle = {
            'available': True,
            'label': SENSECV_OBSTACLE_LABELS[obs_idx] if obs_idx < len(SENSECV_OBSTACLE_LABELS) else str(obs_idx),
            'prob': float(obs_vec[obs_idx]),
            'image': _cam_to_data_uri(obs_cam, base),
        }
        deviation = {
            'available': True,
            'label': SENSECV_DEVIATION_LABELS[dev_idx] if dev_idx < len(SENSECV_DEVIATION_LABELS) else str(dev_idx),
            'prob': float(dev_vec[dev_idx]),
            'image': _cam_to_data_uri(dev_cam, base),
        }
        return obstacle, deviation
    except Exception as e:
        err = {'available': False, 'error': str(e)}
        return err, dict(err)


def activation_maps_frame(idx, time_s, exact=False):
    """Decode the requested frame once and build all three Grad-CAM maps for it."""
    try:
        import cv2
    except Exception as e:
        return {'available': False, 'error': f'OpenCV indisponivel: {e}'}

    video_path = clip_video_path(idx)
    mtime = os.path.getmtime(video_path)
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        return {'available': False, 'error': 'video unreadable'}
    fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    if fps <= 0 or frame_count <= 0:
        cap.release()
        return {'available': False, 'error': 'invalid video metadata'}

    duration = frame_count / fps
    sample_time = max(0.0, min(float(time_s or 0.0), duration))
    if not exact:
        sample_time = math.floor(sample_time * DRONET_SAMPLE_FPS) / DRONET_SAMPLE_FPS
    frame_idx = min(frame_count - 1, max(0, int(round(sample_time * fps))))

    key = (CLIPS[idx], mtime, frame_idx)
    if key in _activation_cache:
        _activation_cache.move_to_end(key)
        cap.release()
        return _activation_cache[key]

    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
    ok, frame = cap.read()
    cap.release()
    if not ok:
        return {'available': False, 'error': 'frame decode failed', 'frame': frame_idx}

    dronet_map = _dronet_activation(frame)
    obstacle_map, deviation_map = _sensemodel_activation(frame)
    result = {
        'available': True,
        'clip': CLIPS[idx],
        'frame': frame_idx,
        'time_s': round(frame_idx / fps, 4),
        'requested_time_s': round(float(time_s or 0.0), 4),
        'sample_fps': DRONET_SAMPLE_FPS,
        'exact': bool(exact),
        'source_fps': round(fps, 4),
        'model': os.path.basename(_sensemodel_runtime['path']),
        'dronet': dronet_map,
        'obstacle': obstacle_map,
        'deviation': deviation_map,
    }
    _activation_cache[key] = result
    if len(_activation_cache) > 24:
        _activation_cache.popitem(last=False)
    return result


# ─── Uploaded-dataset manifest export ────────────────────────────────────────
# Each uploaded dataset (group) may ship an .xlsx manifest describing every clip
# (ID, LOCAL, DESCRICAO, POSICAO CELULAR, LOCAL OBSTACULO, ALTURA OBSTACULO; see
# [[sensecv-02062026-ifce-clip-manifest]]). The manifest export plans a queue of
# every clip whose dataset has a manifest entry, cuts each one with the same
# auto-suggested window as the viewer, and writes per-group outputs under
# data/derived/manifest_exports/<group>/<mode>/ without touching exports/ or
# history.json.
MANIFEST_EXPORTS_DIR = os.path.join(DERIVED_DIR, 'manifest_exports')
_manifest_cache = {}


def _norm_manifest_id(value):
    s = str(value).strip()
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.lower()


def _clip_dataset_root(idx):
    """Top-level uploaded-dataset folder containing this clip, or None."""
    clip_dir = os.path.abspath(clip_path(idx))
    uploads = os.path.abspath(UPLOADS_DIR)
    if os.path.normcase(clip_dir).startswith(os.path.normcase(uploads) + os.sep):
        rel = os.path.relpath(clip_dir, uploads)
        return os.path.join(uploads, rel.split(os.sep)[0])
    return None


def _find_dataset_manifest(dataset_root):
    if not dataset_root or not os.path.isdir(dataset_root):
        return None
    for base, _dirs, files in os.walk(dataset_root):
        for fn in files:
            if fn.lower().endswith('.xlsx') and not fn.startswith('~$'):
                return os.path.join(base, fn)
    return None


def _load_manifest_rows(xlsx_path):
    """Return {normalized_id: {COLUMN: value}} from the first sheet with an ID column."""
    if xlsx_path in _manifest_cache:
        return _manifest_cache[xlsx_path]
    rows = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        header = None
        for raw in ws.iter_rows(values_only=True):
            cells = ['' if v is None else str(v).strip() for v in raw]
            if header is None:
                if any(c.strip().upper() == 'ID' for c in cells):
                    header = [c.strip().upper() for c in cells]
                continue
            record = {header[i]: cells[i] for i in range(min(len(header), len(cells)))}
            cid = record.get('ID', '').strip()
            if cid:
                rows[_norm_manifest_id(cid)] = record
        wb.close()
    except Exception:
        return {}
    _manifest_cache[xlsx_path] = rows
    return rows


def _manifest_label(manifest_id, record):
    parts = [manifest_id]
    for col in ('DESCRICAO', 'POSICAO CELULAR', 'LOCAL OBSTACULO', 'ALTURA OBSTACULO'):
        v = (record.get(col) or '').strip()
        if v:
            parts.append(v)
    return ' | '.join(parts)


def plan_manifest_export(mode='lateral'):
    """Queue of every uploaded clip whose dataset has a matching manifest row."""
    items = []
    for idx, display in enumerate(CLIPS):
        if display.startswith('exports/'):
            continue
        manifest = _find_dataset_manifest(_clip_dataset_root(idx))
        if not manifest:
            continue
        rows = _load_manifest_rows(manifest)
        clip_name = os.path.basename(os.path.normpath(clip_path(idx)))
        record = rows.get(_norm_manifest_id(clip_name))
        if record is None:
            continue
        items.append({
            'clip_idx': idx,
            'source_display': display,
            'group': CLIP_GROUPS.get(display, ''),
            'manifest_id': clip_name,
            'label': _manifest_label(clip_name, record),
        })
    return items


def _manifest_group_dir(group, mode):
    return os.path.join(MANIFEST_EXPORTS_DIR, _safe_clip_name(group or 'sem_grupo'), mode)


def _manifest_clip_out_dir(idx, mode):
    group = CLIP_GROUPS.get(CLIPS[idx], 'sem_grupo')
    clip_name = os.path.basename(os.path.normpath(clip_path(idx)))
    return os.path.join(_manifest_group_dir(group, mode), clip_name)


def export_manifest_clip(idx, mode='lateral'):
    """Cut one uploaded clip into its per-group manifest-export folder.

    Returns a status dict (`ok` / `skipped`) or raises for a true export error
    so the caller can report the exact failing clip.
    """
    import csv
    display = CLIPS[idx]
    group = CLIP_GROUPS.get(display, 'sem_grupo')
    clip_name = os.path.basename(os.path.normpath(clip_path(idx)))
    out_folder = _manifest_clip_out_dir(idx, mode)

    if mode == 'lateral':
        sug = suggest_lateral_deviation(idx)
    else:
        sug = suggest_crop(idx, 'walking')
    if not sug.get('found'):
        return {'status': 'skipped', 'source_display': display,
                'message': str(sug.get('message', 'no_segment'))}

    start, end = float(sug['start']), float(sug['end'])
    if os.path.isdir(out_folder):
        shutil.rmtree(out_folder, ignore_errors=True)
    os.makedirs(out_folder, exist_ok=True)
    out_video = os.path.join(out_folder, clip_name + '.mp4')

    try:
        _ffmpeg_cut(clip_video_path(idx), start, end, out_video)
    except Exception as e:
        shutil.rmtree(out_folder, ignore_errors=True)
        raise RuntimeError(f'ffmpeg_error ({display}): {e}')

    try:
        save_sensor_data(idx, start, end, out_folder)
    except Exception:
        pass

    selection = ssim_frame_selection(idx, start, end)
    try:
        save_ssim_selection(selection, out_folder)
    except Exception:
        pass

    # Append (or refresh) the per-group sources.csv ledger.
    group_dir = _manifest_group_dir(group, mode)
    os.makedirs(group_dir, exist_ok=True)
    csv_path = os.path.join(group_dir, 'sources.csv')
    manifest = _find_dataset_manifest(_clip_dataset_root(idx))
    record = _load_manifest_rows(manifest).get(_norm_manifest_id(clip_name), {}) if manifest else {}
    fieldnames = ['manifest_id', 'output_folder', 'source_display', 'group',
                  'start', 'end', 'duration', 'label']
    existing = []
    if os.path.isfile(csv_path):
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            existing = [r for r in csv.DictReader(f) if r.get('manifest_id') != clip_name]
    existing.append({
        'manifest_id': clip_name,
        'output_folder': clip_name,
        'source_display': display,
        'group': group,
        'start': f'{start:.3f}',
        'end': f'{end:.3f}',
        'duration': f'{end - start:.3f}',
        'label': _manifest_label(clip_name, record),
    })
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(existing)

    return {'status': 'ok', 'source_display': display, 'clip_idx': idx,
            'mode': mode, 'start': round(start, 3), 'end': round(end, 3),
            'output_folder': clip_name, 'video': out_video}


def build_manifest_review(mode='lateral'):
    """One review MP4 of every successful manifest cut, in queue order."""
    videos = []
    for item in plan_manifest_export(mode):
        clip_name = item['manifest_id']
        out_dir = _manifest_clip_out_dir(item['clip_idx'], mode)
        candidate = os.path.join(out_dir, clip_name + '.mp4')
        if os.path.isfile(candidate):
            videos.append(candidate)
    if not videos:
        return {'status': 'error', 'message': 'nenhum clipe exportado encontrado'}
    review_path = os.path.join(MANIFEST_EXPORTS_DIR, f'all_uploaded_{mode}_review.mp4')
    frames = _write_export_review_video(videos, review_path)
    return {'status': 'ok', 'mode': mode, 'videos': len(videos), 'frames': frames,
            'path': review_path, 'url': _derived_file_url(review_path)}


# ─── Revisão: cull bad clips from a group's review video ──────────────────────
# The "Revisão" viewer plays a group's review_all_frames.mp4 (every cut clip's
# frames concatenated at SSIM_REVIEW_FPS) and maps the frame under the playhead
# back to the source clip via a persisted review_index.json. Marking a frame
# "wrong" deletes that whole clip folder from the manifest_exports group (and its
# sources.csv row). The index stays pinned to the *existing* video so the
# frame→clip mapping never shifts mid-review; "Reconstruir" rebuilds the video
# (and a fresh index) from the clips that survived.
REVIEW_VIDEO_NAME = 'review_all_frames.mp4'
REVIEW_INDEX_NAME = 'review_index.json'
# Bump when the index schema changes so stale on-disk indexes get rebuilt.
REVIEW_INDEX_VERSION = 5


def _review_clip_video(group_dir, folder):
    return os.path.join(group_dir, folder, folder + '.mp4')


def _deviation_label(event_type, side):
    """Human label + colour code for a deviation cut's class."""
    if event_type == 'desvio':
        if side == 'LEFT':
            return 'Desvio ◀ esquerda', 'desvio-left'
        if side == 'RIGHT':
            return 'Desvio ▶ direita', 'desvio-right'
        return 'Desvio', 'desvio'
    if event_type == 'parada':
        return 'Parada', 'parada'
    if event_type == 'livre':
        return 'Caminhada livre', 'livre'
    return '—', 'desconhecido'


# The labels the operator can pick in Revisão -> (event_type, side).
REVIEW_LABEL_KINDS = {
    'desvio-left':  ('desvio', 'LEFT'),
    'desvio-right': ('desvio', 'RIGHT'),
    'parada':       ('parada', 'NONE'),
    'livre':        ('livre',  'NONE'),
}
REVIEW_LABELS_NAME = 'review_labels.json'


def _load_label_overrides(group_dir):
    import json
    path = os.path.join(group_dir, REVIEW_LABELS_NAME)
    if os.path.isfile(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_label_overrides(group_dir, data):
    import json
    path = os.path.join(group_dir, REVIEW_LABELS_NAME)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
    except Exception:
        pass


def _read_sources_rows(group_dir):
    """Ordered sources.csv rows (the order export_set wrote = review concat order)."""
    import csv
    path = os.path.join(group_dir, 'sources.csv')
    if not os.path.isfile(path):
        return []
    with open(path, 'r', encoding='utf-8', newline='') as f:
        return list(csv.DictReader(f))


def _video_frame_count(path):
    try:
        import cv2
    except Exception:
        return 0
    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        cap.release()
        return 0
    n = int(round(cap.get(cv2.CAP_PROP_FRAME_COUNT)))
    cap.release()
    return max(0, n)


def build_review_index(group_dir, force=False):
    """Build (or load) the frame→clip map for a group's review video.

    Pinned to review_all_frames.mp4: rebuilt only when missing, forced, or older
    than the video, so exclusions never invalidate offsets already in use.
    Returns the index dict or None when there is no review video yet.
    """
    import json
    video_path = os.path.join(group_dir, REVIEW_VIDEO_NAME)
    if not os.path.isfile(video_path):
        return None
    index_path = os.path.join(group_dir, REVIEW_INDEX_NAME)
    if not force and os.path.isfile(index_path):
        try:
            if os.path.getmtime(index_path) >= os.path.getmtime(video_path):
                with open(index_path, 'r', encoding='utf-8') as f:
                    cached = json.load(f)
                if cached.get('version') == REVIEW_INDEX_VERSION:
                    return cached
        except Exception:
            pass

    # Map each output folder back to its live clip index so we can recover the
    # cut's class (desvio LEFT/RIGHT, parada, livre) from suggest_deviation_cut.
    folder_to_idx = {_safe_clip_name(CLIPS[i]): i for i in range(len(CLIPS))}
    overrides = _load_label_overrides(group_dir)

    clips = []
    cursor = 0
    for row in _read_sources_rows(group_dir):
        if not str(row.get('status', '')).startswith('ok'):
            continue
        folder = row.get('output_folder', '')
        cut = _review_clip_video(group_dir, folder)
        if not folder or not os.path.isfile(cut):
            continue
        n = _video_frame_count(cut)
        if n <= 0:
            continue
        event_type, side = '', ''
        src_idx = folder_to_idx.get(folder)
        # Prefer the class the dataset already carries (an imported cut-dataset
        # ledgers event_type/side in sources.csv or each clip's cut_info.json), so
        # the labels survive even when the source clip isn't loaded. Fall back to
        # re-deriving from the live clip's IMU + manifest.
        event_type = (row.get('event_type', '') or '').strip()
        side = (row.get('side', '') or '').strip()
        if not event_type:
            info = _read_cut_info(group_dir, folder)
            event_type = (info.get('event_type', '') or '').strip()
            side = (info.get('side', '') or '').strip()
        if not event_type and src_idx is not None:
            try:
                sug = suggest_deviation_cut(src_idx)
                event_type = sug.get('event_type', '') or ''
                side = sug.get('side', '') or ''
            except Exception:
                pass
        detected_label, _ = _deviation_label(event_type, side)
        ov = overrides.get(folder)
        overridden = bool(ov)
        if overridden:
            event_type = ov.get('event_type', event_type)
            side = ov.get('side', side)
        label, label_kind = _deviation_label(event_type, side)
        clips.append({
            'folder': folder,
            'source_display': row.get('source_display', ''),
            'source_idx': src_idx,
            'start_frame': cursor,
            'frame_count': n,
            'end_frame': cursor + n,
            'cut_start': row.get('start', ''),
            'cut_end': row.get('end', ''),
            'duration': row.get('duration', ''),
            'status': row.get('status', ''),
            'event_type': event_type,
            'side': side,
            'label': label,
            'label_kind': label_kind,
            'label_overridden': overridden,
            'detected_label': detected_label,
            'excluded': False,
        })
        cursor += n

    index = {
        'version': REVIEW_INDEX_VERSION,
        'fps': SSIM_REVIEW_FPS,
        'total_frames': cursor,
        'video_url': _derived_file_url(video_path),
        'clips': clips,
    }
    try:
        with open(index_path, 'w', encoding='utf-8') as f:
            json.dump(index, f, ensure_ascii=False, indent=1)
    except Exception:
        pass
    return index


def list_review_groups(mode='deviation'):
    """Every manifest-export group that has a review video to cull."""
    out = []
    if not os.path.isdir(MANIFEST_EXPORTS_DIR):
        return out
    for group in sorted(os.listdir(MANIFEST_EXPORTS_DIR)):
        group_dir = os.path.join(MANIFEST_EXPORTS_DIR, group, mode)
        if not os.path.isdir(group_dir):
            continue
        has_review = os.path.isfile(os.path.join(group_dir, REVIEW_VIDEO_NAME))
        n_clips = sum(1 for r in _read_sources_rows(group_dir)
                      if str(r.get('status', '')).startswith('ok'))
        out.append({'group': group, 'mode': mode, 'has_review': has_review,
                    'clips': n_clips})
    return out


def review_exclude_clip(group, mode, folder):
    """Delete one clip folder from a manifest-export group + its sources.csv row.

    Marks the clip excluded in the pinned review_index.json (offsets untouched).
    """
    import csv
    import json
    group_dir = _manifest_group_dir(group, mode)
    base = os.path.abspath(group_dir)
    clip_dir = os.path.abspath(os.path.join(base, folder))
    if not clip_dir.startswith(base + os.sep) or not os.path.basename(clip_dir) == folder:
        return {'status': 'error', 'message': 'folder invalido'}
    if os.path.isdir(clip_dir):
        shutil.rmtree(clip_dir, ignore_errors=True)

    # Drop the row from sources.csv.
    sources = os.path.join(group_dir, 'sources.csv')
    if os.path.isfile(sources):
        rows = _read_sources_rows(group_dir)
        kept = [r for r in rows if r.get('output_folder') != folder]
        if kept and (rows and kept != rows):
            with open(sources, 'w', encoding='utf-8', newline='') as f:
                w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                w.writeheader(); w.writerows(kept)

    # Mark excluded in the pinned index (no offset recompute).
    index_path = os.path.join(group_dir, REVIEW_INDEX_NAME)
    remaining = None
    if os.path.isfile(index_path):
        try:
            with open(index_path, 'r', encoding='utf-8') as f:
                index = json.load(f)
            for c in index.get('clips', []):
                if c.get('folder') == folder:
                    c['excluded'] = True
            remaining = sum(1 for c in index.get('clips', []) if not c.get('excluded'))
            with open(index_path, 'w', encoding='utf-8') as f:
                json.dump(index, f, ensure_ascii=False, indent=1)
        except Exception:
            pass
    return {'status': 'ok', 'group': group, 'mode': mode, 'folder': folder,
            'remaining': remaining}


def review_set_label(group, mode, folder, label_kind):
    """Override one clip's class (desvio LEFT/RIGHT, parada, livre) in Revisão.

    Persisted to review_labels.json so it survives index/video rebuilds, and
    written straight into the pinned review_index.json for the live view.
    """
    import json
    if label_kind not in REVIEW_LABEL_KINDS:
        return {'status': 'error', 'message': f'rotulo invalido: {label_kind}'}
    event_type, side = REVIEW_LABEL_KINDS[label_kind]
    group_dir = _manifest_group_dir(group, mode)
    if not os.path.isdir(os.path.join(group_dir, folder)):
        return {'status': 'error', 'message': 'clipe nao encontrado'}

    overrides = _load_label_overrides(group_dir)
    overrides[folder] = {'event_type': event_type, 'side': side}
    _save_label_overrides(group_dir, overrides)

    label, _ = _deviation_label(event_type, side)
    index_path = os.path.join(group_dir, REVIEW_INDEX_NAME)
    if os.path.isfile(index_path):
        try:
            with open(index_path, 'r', encoding='utf-8') as f:
                index = json.load(f)
            for c in index.get('clips', []):
                if c.get('folder') == folder:
                    c.update(event_type=event_type, side=side, label=label,
                             label_kind=label_kind, label_overridden=True)
            with open(index_path, 'w', encoding='utf-8') as f:
                json.dump(index, f, ensure_ascii=False, indent=1)
        except Exception:
            pass
    return {'status': 'ok', 'group': group, 'mode': mode, 'folder': folder,
            'event_type': event_type, 'side': side, 'label': label,
            'label_kind': label_kind}


def review_rebuild_video(group, mode):
    """Rebuild a group's review video + index from the surviving clips."""
    group_dir = _manifest_group_dir(group, mode)
    videos = []
    for row in _read_sources_rows(group_dir):
        if not str(row.get('status', '')).startswith('ok'):
            continue
        cut = _review_clip_video(group_dir, row.get('output_folder', ''))
        if os.path.isfile(cut):
            videos.append(cut)
    if not videos:
        return {'status': 'error', 'message': 'nenhum clipe restante para revisar'}
    review_path = os.path.join(group_dir, REVIEW_VIDEO_NAME)
    frames = _write_export_review_video(videos, review_path)
    index = build_review_index(group_dir, force=True)
    return {'status': 'ok', 'group': group, 'mode': mode, 'videos': len(videos),
            'frames': frames, 'clips': len((index or {}).get('clips', []))}


def review_recut_clip(group, mode, folder, start, end, clip_idx=None,
                      source_display=None, ssim_threshold=None, ssim_metric='ssim'):
    """Re-cut one review clip to a new [start, end] window in place.

    Re-cuts the clip video and re-saves every sensor sample in the new window
    (no frame selector / SSIM — all frames from the cut are kept), overwriting the
    clip's folder, then updates its sources.csv row. Any manual label override
    survives (it is keyed on the folder name, which does not change). `ssim_*` are
    accepted for backward compatibility with the route and ignored.
    """
    import csv
    group_dir = _manifest_group_dir(group, mode)
    base = os.path.abspath(group_dir)
    out_folder = os.path.abspath(os.path.join(base, folder))
    if not out_folder.startswith(base + os.sep) or os.path.basename(out_folder) != folder:
        return {'status': 'error', 'message': 'folder invalido'}
    if not os.path.isdir(group_dir):
        return {'status': 'error',
                'message': f'grupo "{group}" não encontrado (recarregue a revisão)'}

    rows = _read_sources_rows(group_dir)
    row = next((r for r in rows if r.get('output_folder') == folder), None)
    if row is None:
        # Self-heal: the ledger lost this clip's row (or was rewritten by another
        # operation). Rebuild a minimal row matching the existing schema so the
        # recut still lands instead of failing. The caller's source_display /
        # clip_idx still resolve the source clip below.
        fieldnames = list(rows[0].keys()) if rows else [
            'output_folder', 'source_display', 'start', 'end', 'duration',
            'event_type', 'side', 'status']
        row = {k: '' for k in fieldnames}
        row['output_folder'] = folder
        if 'source_display' in row and source_display:
            row['source_display'] = source_display
        if 'status' in row:
            row['status'] = 'ok'
        rows.append(row)

    # Resolve the live source-clip index: the sources.csv display name wins; the
    # caller's source_display / clip_idx are fallbacks when CLIPS ordering shifted
    # or the ledger row was just rebuilt.
    idx = None
    src_display = row.get('source_display', '') or source_display or ''
    if src_display:
        try:
            idx = CLIPS.index(src_display)
        except ValueError:
            idx = None
    if idx is None and clip_idx is not None:
        try:
            ci = int(clip_idx)
            if 0 <= ci < len(CLIPS):
                idx = ci
        except (TypeError, ValueError):
            idx = None
    if idx is None:
        return {'status': 'error', 'message': 'clipe de origem nao encontrado'}

    try:
        start, end = float(start), float(end)
    except (TypeError, ValueError):
        return {'status': 'error', 'message': 'janela invalida'}
    if not (end > start):
        return {'status': 'error', 'message': 'fim deve ser maior que inicio'}

    out_video = os.path.join(out_folder, folder + '.mp4')
    if os.path.isdir(out_folder):
        shutil.rmtree(out_folder, ignore_errors=True)
    os.makedirs(out_folder)
    try:
        _ffmpeg_cut(clip_video_path(idx), start, end, out_video)
    except Exception as e:
        shutil.rmtree(out_folder, ignore_errors=True)
        return {'status': 'error', 'message': f'ffmpeg_error: {e}'}

    sensor_error = None
    try:
        save_sensor_data(idx, start, end, out_folder)
    except Exception as e:
        sensor_error = str(e)

    # Update only columns common to every sources.csv schema (export and uploaded
    # cut-datasets alike); never add ssim columns the uploaded ledger lacks.
    row.update(
        start=f'{start:.3f}', end=f'{end:.3f}', duration=f'{end - start:.3f}',
        status=('ok_no_sensor: ' + sensor_error) if sensor_error else 'ok',
    )
    if rows:
        with open(os.path.join(group_dir, 'sources.csv'), 'w', encoding='utf-8', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader(); w.writerows(rows)

    # Patch the pinned index in place instead of re-encoding the whole review
    # video (that concat is the slow part). The new cut is on disk + in
    # sources.csv; we flag this clip pending so the UI shows the cut was saved but
    # is not in review_all_frames.mp4 yet. Offsets/frame_count stay put — the
    # existing video is untouched until the operator clicks "Reconstruir vídeo".
    import json
    index_path = os.path.join(group_dir, REVIEW_INDEX_NAME)
    if os.path.isfile(index_path):
        try:
            with open(index_path, 'r', encoding='utf-8') as f:
                index = json.load(f)
            for c in index.get('clips', []):
                if c.get('folder') == folder:
                    c['cut_start'] = f'{start:.3f}'
                    c['cut_end'] = f'{end:.3f}'
                    c['duration'] = f'{end - start:.3f}'
                    c['recut_pending'] = True
            with open(index_path, 'w', encoding='utf-8') as f:
                json.dump(index, f, ensure_ascii=False, indent=1)
        except Exception:
            pass

    return {'status': 'ok', 'group': group, 'mode': mode, 'folder': folder,
            'start': f'{start:.3f}', 'end': f'{end:.3f}',
            'duration': f'{end - start:.3f}', 'recut_pending': True}


# ─── Revisão: import an external cut-dataset (e.g. datasetcortado) ────────────
# A cut-dataset is a folder of per-clip subfolders, each holding the trimmed
# clip video (and optionally cut_info.json with its class). `datasetcortado`
# (scripts/make_datasetcortado.py) is the canonical example. Uploading a .zip of
# one copies the clip videos into a manifest_exports review group and builds the
# review video + index, so the existing Revisão tools (exclude / relabel / recut
# / rebuild) work on it unchanged.


def _read_cut_info(group_dir, folder):
    """A clip's cut_info.json (class/window), if the dataset shipped one."""
    import json
    path = os.path.join(group_dir, folder, 'cut_info.json')
    if os.path.isfile(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _iter_cut_clips(root):
    """Yield (folder_name, clip_dir, video_path) for subfolders that hold a video."""
    if not os.path.isdir(root):
        return
    for name in sorted(os.listdir(root)):
        clip_dir = os.path.join(root, name)
        if not os.path.isdir(clip_dir):
            continue
        video = _clip_video_file(clip_dir)
        if video:
            yield name, clip_dir, video


def _find_cut_dataset_root(base):
    """Locate the directory that actually holds the clip subfolders.

    An uploaded zip may wrap the dataset in one or more parent folders; pick the
    directory with the most clip subfolders (ties broken by a sources.csv).
    """
    best, best_score = None, 0
    for cur, _dirs, files in os.walk(base):
        n = sum(1 for _ in _iter_cut_clips(cur))
        if n == 0 and 'sources.csv' not in files:
            continue
        score = n * 2 + (1 if 'sources.csv' in files else 0)
        if score > best_score:
            best, best_score = cur, score
    return best


def ingest_review_dataset(src_dir, group, mode='deviation'):
    """Copy a cut-dataset into a review group and build its review video + index.

    Only the clip videos (renamed to <folder>/<folder>.mp4) and any cut_info.json
    are copied — that is all the review needs. A normalized sources.csv is written
    from the dataset's own ledger (or synthesized from the folders), then the
    review video and frame→clip index are built.
    """
    import csv
    import json
    clips = list(_iter_cut_clips(src_dir))
    if not clips:
        return {'status': 'error',
                'message': 'dataset sem clipes (esperado subpastas com .mp4)'}

    group_dir = _manifest_group_dir(group, mode)
    if os.path.isdir(group_dir):
        shutil.rmtree(group_dir, ignore_errors=True)
    os.makedirs(group_dir, exist_ok=True)

    # Pull any per-clip labels the dataset already carries.
    src_rows = {}
    src_csv = os.path.join(src_dir, 'sources.csv')
    if os.path.isfile(src_csv):
        try:
            with open(src_csv, 'r', encoding='utf-8', newline='') as f:
                for r in csv.DictReader(f):
                    if r.get('output_folder'):
                        src_rows[r['output_folder']] = r
        except Exception:
            src_rows = {}

    fields = ['output_folder', 'source_display', 'start', 'end', 'duration',
              'event_type', 'side', 'status']
    rows = []
    for folder, clip_dir, video in clips:
        dst = os.path.join(group_dir, folder)
        os.makedirs(dst, exist_ok=True)
        shutil.copy2(video, os.path.join(dst, folder + '.mp4'))
        info = {}
        ci = os.path.join(clip_dir, 'cut_info.json')
        if os.path.isfile(ci):
            try:
                with open(ci, 'r', encoding='utf-8') as f:
                    info = json.load(f)
            except Exception:
                info = {}
            try:
                shutil.copy2(ci, os.path.join(dst, 'cut_info.json'))
            except Exception:
                pass
        sr = src_rows.get(folder, {})
        rows.append({
            'output_folder': folder,
            'source_display': sr.get('source_display') or info.get('source_display') or folder,
            'start': sr.get('start') or info.get('start', ''),
            'end': sr.get('end') or info.get('end', ''),
            'duration': sr.get('duration') or info.get('duration', ''),
            'event_type': sr.get('event_type') or info.get('event_type', ''),
            'side': sr.get('side') or info.get('side', ''),
            'status': sr.get('status') or 'ok',
        })

    with open(os.path.join(group_dir, 'sources.csv'), 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader(); w.writerows(rows)

    rebuilt = review_rebuild_video(group, mode)
    if rebuilt.get('status') != 'ok':
        return {'status': 'error',
                'message': rebuilt.get('message', 'falha ao montar o vídeo de revisão')}
    return {'status': 'ok', 'group': _safe_clip_name(group), 'mode': mode,
            'clips': len(rows), 'frames': rebuilt.get('frames')}


# ─── Deviation inspection video ──────────────────────────────────────────────
# Two outputs from one request (no ML, sensor-only):
#   1. Real exports: each deviation clip's auto-suggested lateral cut, exported
#      with the SAME pipeline as a common export (clean cut + sensors + ssim, via
#      export_set). No overlay is burned into these files.
#   2. One validation video: every exported cut's frames concatenated at 30 fps
#      (no title cards / transitions), each carrying a "Desvio: LEFT|RIGHT"
#      header banner. The header lives ONLY in this validation video.
# Only clips with a detected deviation header (LEFT/RIGHT) are exported; clips
# the sensors read as NONE are skipped.
INSPECT_PLAY_FPS = float(os.environ.get('SENSECV_INSPECT_FPS', '30'))
DEVIATION_EXPORTS_DIR = os.path.join(DERIVED_DIR, 'deviation_exports')
VALIDATION_VIDEO_PATH = os.path.join(DEVIATION_EXPORTS_DIR, 'validation.mp4')
# BGR colours for cv2 overlays.
_DEV_COLORS = {'LEFT': (80, 170, 255), 'RIGHT': (80, 170, 255), 'NONE': (120, 230, 120)}


def _draw_label_bar(frame, lines):
    """Draw a translucent top bar with one coloured (text, BGR) line per entry."""
    import cv2
    h, w = frame.shape[:2]
    fs = max(0.5, w / 1100.0)
    th = max(1, int(round(fs * 2)))
    line_h = int(40 * fs)
    pad = int(10 * fs)
    bar_h = pad * 2 + line_h * len(lines)
    overlay = frame.copy()
    cv2.rectangle(overlay, (0, 0), (w, bar_h), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.55, frame, 0.45, 0, frame)
    y = pad + int(line_h * 0.72)
    for text, col in lines:
        cv2.putText(frame, text, (pad + 4, y), cv2.FONT_HERSHEY_SIMPLEX, fs,
                    col, th, cv2.LINE_AA)
        y += line_h
    return frame


def _write_validation_video(clips, out_video, fps):
    """Concatenate the exported cut videos into one validation MP4.

    `clips` is a list of dicts with `video_path` and `deviation` (LEFT/RIGHT).
    Every source frame is written back-to-back at a constant `fps` (no resampling
    to real time — playback runs faster than wall-clock, by design) with a
    "Desvio: <side>" header burned in. No title cards / transitions. The header
    exists only here; the exported clips themselves stay clean.
    """
    import cv2
    try:
        import imageio_ffmpeg
        ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    except Exception as e:
        raise RuntimeError(f'ffmpeg indisponivel para video de validacao: {e}')

    clips = [c for c in clips if c.get('video_path') and os.path.isfile(c['video_path'])]
    if not clips:
        return 0

    width = height = None
    for c in clips:
        cap = cv2.VideoCapture(c['video_path'])
        if cap.isOpened():
            width, height = _review_frame_size(
                cap.get(cv2.CAP_PROP_FRAME_WIDTH), cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            cap.release()
            break
        cap.release()
    if width is None:
        return 0

    os.makedirs(os.path.dirname(out_video), exist_ok=True)
    cmd = [
        ffmpeg, '-y', '-f', 'rawvideo', '-vcodec', 'rawvideo', '-pix_fmt', 'rgb24',
        '-s', f'{width}x{height}', '-r', f'{fps}', '-i', '-', '-an',
        '-c:v', 'libx264', '-preset', 'ultrafast', '-crf', '23',
        '-pix_fmt', 'yuv420p', '-movflags', '+faststart', out_video,
    ]
    proc = subprocess.Popen(cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    written = 0
    try:
        for c in clips:
            side = c['deviation']
            dev_col = _DEV_COLORS.get(side, (235, 235, 235))
            label = c.get('clip', '')
            cap = cv2.VideoCapture(c['video_path'])
            if not cap.isOpened():
                cap.release()
                continue
            while True:
                ok, frame = cap.read()
                if not ok:
                    break
                if frame.shape[1] != width or frame.shape[0] != height:
                    frame = cv2.resize(frame, (width, height), interpolation=cv2.INTER_AREA)
                _draw_label_bar(frame, [
                    (label, (235, 235, 235)),
                    (f'Desvio: {side}', dev_col),
                ])
                proc.stdin.write(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB).tobytes())
                written += 1
            cap.release()
    finally:
        if proc.stdin:
            proc.stdin.close()
    stderr = proc.stderr.read().decode('utf-8', errors='replace')
    proc.wait()
    if proc.returncode != 0:
        raise RuntimeError(stderr[-500:])
    return written


def export_deviation_set(indices, play_fps=None):
    """Export every clip whose sensors read a deviation, then build one
    validation video over those exports.

    Steps:
      1. For each clip, take its PDF-based deviation cut (`suggest_deviation_cut`:
         the T1-Δ→T1 decision window by default) and the event's direction. Keep
         only LEFT/RIGHT — clips with no IMU desvio are skipped ("only the videos
         with the equivalent header are exported").
      2. Export the kept clips with `export_set` (mode='deviation'): the exact
         same clean cut + sensors + ssim a common export produces, no overlay.
      3. Concatenate the exported cuts into a 30 fps validation video with the
         deviation header burned in (header lives only in that video).
    """
    play_fps = float(play_fps or INSPECT_PLAY_FPS) or INSPECT_PLAY_FPS

    # Phase 1 — pick deviation clips (snapshot idx-based access up front). The cut
    # window and side both come from the same IMU event, so they stay consistent.
    plan = []
    for idx in indices:
        try:
            display = CLIPS[idx]
        except (KeyError, IndexError):
            continue
        cut = suggest_deviation_cut(idx)
        if not cut.get('found') or cut.get('side') not in ('LEFT', 'RIGHT'):
            plan.append({'clip': display, 'status': 'skipped_no_deviation',
                         'deviation': 'NONE'})
            continue
        plan.append({'idx': idx, 'clip': display, 'status': 'planned',
                     'start': cut['start'], 'end': cut['end'], 'deviation': cut['side'],
                     't1': cut['t1'], 't2': cut['t2'], 'window': cut['window'],
                     'confidence': cut['confidence']})

    keep = [p for p in plan if p['status'] == 'planned']
    if not keep:
        raise RuntimeError('nenhum clipe com desvio (LEFT/RIGHT) para exportar')

    # Phase 2 — real exports (same type as a common export). No review here; the
    # validation video below replaces it. export_set re-runs suggest_deviation_cut
    # so the exported window matches the plan.
    summary = export_set([p['idx'] for p in keep], DEVIATION_EXPORTS_DIR,
                         mode='deviation', build_review=False)

    # Resolve each kept clip's exported cut path (export_set names by clip).
    for p in keep:
        name = _safe_clip_name(p['clip'])
        vp = os.path.join(DEVIATION_EXPORTS_DIR, name, name + '.mp4')
        if os.path.isfile(vp):
            p['video_path'] = vp
            p['status'] = 'exported'
            p['export_folder'] = name
        else:
            p['status'] = 'export_failed'

    exported = [p for p in keep if p['status'] == 'exported']
    val_frames = _write_validation_video(exported, VALIDATION_VIDEO_PATH, play_fps)

    counts = {'LEFT': 0, 'RIGHT': 0}
    for p in exported:
        counts[p['deviation']] = counts.get(p['deviation'], 0) + 1
    return {
        'status': 'ok',
        'requested': len(indices),
        'exported': len(exported),
        'skipped_none': sum(1 for p in plan if p['status'] == 'skipped_no_deviation'),
        'counts': counts,
        'out_dir': DEVIATION_EXPORTS_DIR,
        'csv_url': summary.get('csv_url'),
        'validation_url': _derived_file_url(VALIDATION_VIDEO_PATH) if val_frames else None,
        'validation_frames': val_frames,
        'play_fps': play_fps,
        'clips': [{k: v for k, v in p.items() if k not in ('video_path', 'idx')}
                  for p in plan],
    }


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    refresh_clips()
    h = load_history()
    return render_template('index.html',
                           total=len(CLIPS),
                           clips=json.dumps(CLIPS),
                           clip_groups=json.dumps(clip_groups()),
                           history=json.dumps(h),
                           export_folders=json.dumps(list_export_folders()),
                           next_number=get_next_number(),
                           ssim_threshold=SSIM_THRESHOLD)

@app.route('/video/<int:idx>')
def video(idx):
    if idx<0 or idx>=len(CLIPS): return 'Not found',404
    path = clip_video_path(idx); size=os.path.getsize(path)
    rng  = request.headers.get('Range')
    if not rng:
        with open(path,'rb') as f: data=f.read()
        return Response(data,mimetype='video/mp4',
                        headers={'Accept-Ranges':'bytes','Content-Length':str(size)})
    m=re.search(r'(\d+)-(\d*)',rng); b1=int(m.group(1)); b2=int(m.group(2)) if m.group(2) else size-1
    length=b2-b1+1
    with open(path,'rb') as f: f.seek(b1); data=f.read(length)
    return Response(data,206,mimetype='video/mp4',headers={
        'Content-Range':f'bytes {b1}-{b2}/{size}','Accept-Ranges':'bytes','Content-Length':str(length)})

@app.route('/frame/<int:idx>/<int:source_index>.jpg')
def frame_jpg(idx, source_index):
    if idx<0 or idx>=len(CLIPS): return 'Not found',404
    frames = load_json(clip_path(idx), 'frames.json')['frames']
    if source_index < 0 or source_index >= len(frames):
        return 'Not found',404
    try:
        import cv2
        cap = cv2.VideoCapture(clip_video_path(idx))
    except Exception as e:
        return str(e),503
    if not cap.isOpened():
        return 'video unreadable',503
    video_fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
    if video_fps <= 0:
        cap.release()
        return 'invalid video fps',503
    t0 = float(frames[0]['time_usec'])
    time_s = (float(frames[source_index]['time_usec']) - t0) / 1e6
    frame_no = max(0, int(round(time_s * video_fps)))
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_no)
    ok, frame = cap.read()
    cap.release()
    if not ok:
        return 'frame decode failed',503
    h, w = frame.shape[:2]
    max_w = 360
    if w > max_w:
        scale = max_w / float(w)
        frame = cv2.resize(frame, (max_w, max(1, int(round(h * scale)))),
                           interpolation=cv2.INTER_AREA)
    ok, encoded = cv2.imencode('.jpg', frame, [int(cv2.IMWRITE_JPEG_QUALITY), 82])
    if not ok:
        return 'jpeg encode failed',503
    return Response(encoded.tobytes(), mimetype='image/jpeg')

@app.route('/export-file/<folder>/<path:relpath>')
def export_file(folder, relpath):
    if folder not in list_export_folders():
        return 'Not found',404
    base = os.path.abspath(os.path.join(EXPORTS_DIR, folder))
    target = os.path.abspath(os.path.join(base, relpath))
    if not target.startswith(base + os.sep) or not os.path.isfile(target):
        return 'Not found',404
    return send_from_directory(base, relpath, conditional=True)


@app.route('/derived-file/<path:relpath>')
def derived_file(relpath):
    base = os.path.abspath(DERIVED_DIR)
    target = os.path.abspath(os.path.join(base, relpath))
    if not target.startswith(base + os.sep) or not os.path.isfile(target):
        return 'Not found',404
    return send_from_directory(base, relpath, conditional=True)

@app.route('/api/data/<int:idx>')
def api_data(idx):
    if idx<0 or idx>=len(CLIPS): return jsonify({'error':'out of range'}),404
    return jsonify(get_clip_data(idx))

@app.route('/api/dronet/<int:idx>')
def api_dronet(idx):
    if idx<0 or idx>=len(CLIPS): return jsonify({'available':False,'error':'out of range'}),404
    try:
        time_s = float(request.args.get('time', '0') or 0)
    except ValueError:
        time_s = 0.0
    exact = request.args.get('exact') in ('1', 'true', 'yes')
    result = dronet_frame_classification(idx, time_s, exact=exact)
    status = 200 if result.get('available') else 503
    return jsonify(result), status

@app.route('/api/sensemodel/<int:idx>')
def api_sensemodel(idx):
    if idx<0 or idx>=len(CLIPS): return jsonify({'available':False,'error':'out of range'}),404
    try:
        time_s = float(request.args.get('time', '0') or 0)
    except ValueError:
        time_s = 0.0
    exact = request.args.get('exact') in ('1', 'true', 'yes')
    result = sensemodel_frame_classification(idx, time_s, exact=exact)
    status = 200 if result.get('available') else 503
    return jsonify(result), status

@app.route('/api/activation/<int:idx>')
def api_activation(idx):
    if idx<0 or idx>=len(CLIPS): return jsonify({'available':False,'error':'out of range'}),404
    try:
        time_s = float(request.args.get('time', '0') or 0)
    except ValueError:
        time_s = 0.0
    exact = request.args.get('exact') in ('1', 'true', 'yes')
    result = activation_maps_frame(idx, time_s, exact=exact)
    status = 200 if result.get('available') else 503
    return jsonify(result), status

def _sensemodel_info():
    path = _sensemodel_runtime['path']
    return {
        'name': os.path.basename(path),
        'path': path,
        'exists': os.path.isfile(path),
        'is_default': os.path.abspath(path) == os.path.abspath(SENSECV_MODEL_PATH),
        'default_name': os.path.basename(SENSECV_MODEL_PATH),
        'loaded': _sensemodel_runtime['model'] is not None,
        'error': _sensemodel_runtime['error'],
    }

@app.route('/api/sensemodel-info')
def api_sensemodel_info():
    return jsonify(_sensemodel_info())

@app.route('/api/upload-model', methods=['POST'])
def api_upload_model():
    upload = request.files.get('model') or request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'status':'error','message':'Arquivo .keras ausente'}),400
    if not upload.filename.lower().endswith('.keras'):
        return jsonify({'status':'error','message':'Envie um arquivo .keras'}),400

    # _safe_name strips the extension; re-add it so the saved file stays loadable.
    target = os.path.join(MODELS_DIR, _safe_name(upload.filename) + '.keras')
    try:
        upload.save(target)
    except Exception as e:
        return jsonify({'status':'error','message':f'falha ao salvar modelo: {e}'}),500

    # Activate the uploaded model and force a load now so the operator gets an
    # immediate, actionable result (e.g. an incompatible-architecture error)
    # instead of a silent failure on the first inference request.
    _set_sensemodel_path(target)
    runtime = _load_sensemodel_runtime()
    info = _sensemodel_info()
    if runtime['error']:
        return jsonify({
            'status':'error',
            'message':f'Modelo salvo, mas falhou ao carregar: {runtime["error"]}',
            **info,
        }),400
    return jsonify({
        'status':'ok',
        'message':f'Modelo {info["name"]} carregado',
        **info,
    })

@app.route('/api/inspect-deviation', methods=['POST'])
def api_inspect_deviation():
    refresh_clips()
    if not CLIPS:
        return jsonify({'status':'error','message':'nenhum clipe disponivel'}),400
    body = request.get_json(silent=True) or {}
    count = body.get('count', 'all')
    # 'all'/0/negative -> every clip; otherwise the first N clips.
    if isinstance(count, str) and count.strip().lower() in ('all', 'todos', ''):
        indices = list(range(len(CLIPS)))
    else:
        try:
            n = int(count)
        except (TypeError, ValueError):
            return jsonify({'status':'error','message':'count invalido'}),400
        indices = list(range(len(CLIPS))) if n <= 0 else list(range(min(n, len(CLIPS))))

    try:
        play_fps = float(body.get('play_fps') or INSPECT_PLAY_FPS)
    except (TypeError, ValueError):
        play_fps = INSPECT_PLAY_FPS

    try:
        result = export_deviation_set(indices, play_fps=play_fps)
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}),503
    return jsonify(result)

@app.route('/api/reset-model', methods=['POST'])
def api_reset_model():
    _set_sensemodel_path(SENSECV_MODEL_PATH)
    runtime = _load_sensemodel_runtime()
    info = _sensemodel_info()
    if runtime['error']:
        return jsonify({'status':'error','message':runtime['error'], **info}),400
    return jsonify({
        'status':'ok',
        'message':f'Modelo padrão restaurado ({info["name"]})',
        **info,
    })

@app.route('/api/clips')
def api_clips():
    refresh_clips()
    return jsonify({'clips':CLIPS,'groups':clip_groups(),'total':len(CLIPS)})

@app.route('/api/upload-zip', methods=['POST'])
def api_upload_zip():
    upload = request.files.get('zip') or request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'status':'error','message':'Arquivo .zip ausente'}),400
    if not upload.filename.lower().endswith('.zip'):
        return jsonify({'status':'error','message':'Envie um arquivo .zip'}),400

    target = _unique_dir(UPLOADS_DIR, upload.filename)
    try:
        with zipfile.ZipFile(upload.stream) as zf:
            bad = zf.testzip()
            if bad:
                raise ValueError(f'arquivo corrompido dentro do zip: {bad}')
            _safe_extract_zip(zf, target)
        clip_dirs = _discover_clip_dirs(target, recursive=True)
        if not clip_dirs:
            shutil.rmtree(target, ignore_errors=True)
            return jsonify({
                'status':'error',
                'message':'O zip nao contem subpastas SenseCV validas com video.mp4 e frames.json'
            }),400
    except zipfile.BadZipFile:
        shutil.rmtree(target, ignore_errors=True)
        return jsonify({'status':'error','message':'Zip invalido ou corrompido'}),400
    except Exception as e:
        shutil.rmtree(target, ignore_errors=True)
        return jsonify({'status':'error','message':str(e)}),500

    refresh_clips()
    rel = os.path.relpath(target, UPLOADS_DIR).replace(os.sep, '/')
    added = [
        name for name, path in CLIP_PATHS.items()
        if os.path.commonpath([os.path.abspath(target), os.path.abspath(path)]) == os.path.abspath(target)
    ]
    return jsonify({
        'status':'ok',
        'dataset': rel,
        'clips_added': len(added),
        'clips': CLIPS,
        'groups': clip_groups(),
        'total': len(CLIPS),
    })

@app.route('/api/history')
def api_history():
    return jsonify(load_history())

@app.route('/api/export-state')
def api_export_state():
    refresh_clips()
    return jsonify({
        'clips': CLIPS,
        'groups': clip_groups(),
        'total': len(CLIPS),
        'history': load_history(),
        'export_folders': list_export_folders(),
        'next_number': get_next_number(),
    })

@app.route('/api/next-number')
def api_next_number():
    return jsonify({'number': get_next_number()})

@app.route('/api/suggest/<int:idx>')
def api_suggest(idx):
    if idx<0 or idx>=len(CLIPS): return jsonify({'found':False}),404
    mode = request.args.get('mode', 'walking')
    try:
        metric = _request_metric()
        threshold = _request_ssim_threshold(metric=metric)
    except ValueError as e:
        return jsonify({'found':False, 'error':str(e)}),400
    if mode == 'lateral':
        # Viewer "Desvio lateral" button: use the PDF-based deviation cut
        # (decisão window T1-Δ→T1 by default), not the old lateral-velocity
        # heuristic. Returns found:false when the IMU finds no desvio, so
        # runAutoSuggest() falls back to walking for those clips.
        return jsonify(_suggestion_with_ssim(idx, suggest_deviation_cut(idx), threshold=threshold, metric=metric))
    # Only 'walking' remains (the old 'vertical' suggestion was retired).
    return jsonify(_suggestion_with_ssim(idx, suggest_crop(idx, 'walking'), threshold=threshold, metric=metric))

@app.route('/api/ssim/<int:idx>')
def api_ssim(idx):
    if idx<0 or idx>=len(CLIPS): return jsonify({'error':'out of range'}),404
    try:
        start = float(request.args.get('start', '0') or 0)
        end = float(request.args.get('end', '0') or 0)
        metric = _request_metric()
        threshold = _request_ssim_threshold(metric=metric)
    except ValueError as e:
        if 'SSIM' in str(e) or 'metric' in str(e):
            return jsonify({'error':str(e)}),400
        return jsonify({'error':'invalid start/end'}),400
    if end <= start:
        return jsonify({'error':'end must be greater than start'}),400
    try:
        return jsonify(ssim_selection_payload(idx, start, end, threshold=threshold, metric=metric))
    except Exception as e:
        return jsonify({'error':str(e)}),500

@app.route('/api/imu-events/<int:idx>')
def api_imu_events(idx):
    if idx < 0 or idx >= len(CLIPS):
        return jsonify({'error': 'out of range'}), 404
    try:
        delta = float(request.args.get('delta') or IMU_EVENT_DELTA_SEC)
        if not 0.0 < delta <= 5.0:
            raise ValueError
    except ValueError:
        return jsonify({'error': 'delta must be between 0 and 5 seconds'}), 400
    try:
        return jsonify(detect_imu_events(idx, delta=delta))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/batch-export', methods=['POST'])
def api_batch_export():
    body  = request.json or {}
    preset = body.get('preset', 'sensecv')
    mode   = body.get('mode', 'walking')
    if preset not in PRESETS or mode not in ('walking', 'lateral'):
        return jsonify({'status':'error','message':'invalid preset/mode'}),400
    flt = _preset_filter(preset)
    indices = [i for i, n in enumerate(CLIPS) if flt(n)]
    out_dir = _preset_out_dir(preset, mode)
    summary = export_set(indices, out_dir, mode=mode)
    return jsonify({'status':'ok','preset':preset,'mode':mode, **summary})

@app.route('/api/manifest-export', methods=['POST'])
def api_manifest_export():
    body = request.json or {}
    mode = body.get('mode', 'lateral')
    if mode not in ('walking', 'lateral'):
        return jsonify({'status':'error','message':'invalid mode'}),400
    reset = body.get('reset', True)
    if reset and os.path.isdir(MANIFEST_EXPORTS_DIR):
        shutil.rmtree(MANIFEST_EXPORTS_DIR, ignore_errors=True)
    items = plan_manifest_export(mode)
    return jsonify({'status':'ok','mode':mode,'total':len(items),'items':items})

@app.route('/api/manifest-export/clip', methods=['POST'])
def api_manifest_export_clip():
    body = request.json or {}
    mode = body.get('mode', 'lateral')
    if mode not in ('walking', 'lateral'):
        return jsonify({'status':'error','message':'invalid mode'}),400
    try:
        idx = int(body.get('clip_idx'))
    except (TypeError, ValueError):
        return jsonify({'status':'error','message':'clip_idx invalido'}),400
    if idx<0 or idx>=len(CLIPS):
        return jsonify({'status':'error','message':'out of range'}),404
    try:
        return jsonify(export_manifest_clip(idx, mode=mode))
    except Exception as e:
        return jsonify({'status':'error','source_display':CLIPS[idx],'message':str(e)}),500

@app.route('/api/manifest-export/review', methods=['POST'])
def api_manifest_export_review():
    body = request.json or {}
    mode = body.get('mode', 'lateral')
    if mode not in ('walking', 'lateral'):
        return jsonify({'status':'error','message':'invalid mode'}),400
    try:
        result = build_manifest_review(mode)
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}),500
    status = 200 if result.get('status') == 'ok' else 500
    return jsonify(result), status

@app.route('/api/revisao/groups')
def api_revisao_groups():
    mode = request.args.get('mode', 'deviation')
    return jsonify({'status': 'ok', 'mode': mode, 'groups': list_review_groups(mode)})

@app.route('/api/revisao/index')
def api_revisao_index():
    mode = request.args.get('mode', 'deviation')
    group = request.args.get('group', '')
    if not group:
        return jsonify({'status': 'error', 'message': 'group obrigatorio'}), 400
    group_dir = _manifest_group_dir(group, mode)
    index = build_review_index(group_dir)
    if index is None:
        return jsonify({'status': 'error', 'message': 'sem video de revisao para este grupo'}), 404
    return jsonify({'status': 'ok', 'group': group, 'mode': mode, **index})

@app.route('/api/revisao/exclude', methods=['POST'])
def api_revisao_exclude():
    body = request.json or {}
    group = body.get('group', '')
    mode = body.get('mode', 'deviation')
    folder = body.get('folder', '')
    if not group or not folder:
        return jsonify({'status': 'error', 'message': 'group e folder obrigatorios'}), 400
    result = review_exclude_clip(group, mode, folder)
    return jsonify(result), (200 if result.get('status') == 'ok' else 400)

@app.route('/api/revisao/label', methods=['POST'])
def api_revisao_label():
    body = request.json or {}
    group = body.get('group', '')
    mode = body.get('mode', 'deviation')
    folder = body.get('folder', '')
    label_kind = body.get('label_kind', '')
    if not group or not folder:
        return jsonify({'status': 'error', 'message': 'group e folder obrigatorios'}), 400
    result = review_set_label(group, mode, folder, label_kind)
    return jsonify(result), (200 if result.get('status') == 'ok' else 400)

@app.route('/api/revisao/rebuild', methods=['POST'])
def api_revisao_rebuild():
    body = request.json or {}
    group = body.get('group', '')
    mode = body.get('mode', 'deviation')
    if not group:
        return jsonify({'status': 'error', 'message': 'group obrigatorio'}), 400
    try:
        result = review_rebuild_video(group, mode)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    return jsonify(result), (200 if result.get('status') == 'ok' else 400)

@app.route('/api/revisao/recut', methods=['POST'])
def api_revisao_recut():
    body = request.json or {}
    group = body.get('group', '')
    mode = body.get('mode', 'deviation')
    folder = body.get('folder', '')
    if not group or not folder:
        return jsonify({'status': 'error', 'message': 'group e folder obrigatorios'}), 400
    if body.get('start') is None or body.get('end') is None:
        return jsonify({'status': 'error', 'message': 'start e end obrigatorios'}), 400
    try:
        ssim_metric = _coerce_metric(body.get('ssim_metric'))
        ssim_threshold = _body_ssim_threshold(body, metric=ssim_metric)
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400
    try:
        result = review_recut_clip(
            group, mode, folder, body.get('start'), body.get('end'),
            clip_idx=body.get('clip_idx'), source_display=body.get('source_display'),
            ssim_threshold=ssim_threshold, ssim_metric=ssim_metric)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    return jsonify(result), (200 if result.get('status') == 'ok' else 400)

@app.route('/api/revisao/upload', methods=['POST'])
def api_revisao_upload():
    upload = request.files.get('zip') or request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'status': 'error', 'message': 'Arquivo .zip ausente'}), 400
    if not upload.filename.lower().endswith('.zip'):
        return jsonify({'status': 'error', 'message': 'Envie um arquivo .zip'}), 400
    group = (request.form.get('name') or '').strip() \
        or os.path.splitext(os.path.basename(upload.filename))[0]
    tmp = _unique_dir(os.path.join(DERIVED_DIR, 'review_uploads'), upload.filename)
    try:
        with zipfile.ZipFile(upload.stream) as zf:
            bad = zf.testzip()
            if bad:
                raise ValueError(f'arquivo corrompido dentro do zip: {bad}')
            _safe_extract_zip(zf, tmp)
        root = _find_cut_dataset_root(tmp)
        if not root:
            return jsonify({'status': 'error',
                            'message': 'o zip nao contem subpastas de clipes com .mp4'}), 400
        result = ingest_review_dataset(root, group)
    except zipfile.BadZipFile:
        return jsonify({'status': 'error', 'message': 'Zip invalido ou corrompido'}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    return jsonify(result), (200 if result.get('status') == 'ok' else 400)

@app.route('/api/crop', methods=['POST'])
def api_crop():
    body  = request.json
    idx   = int(body.get('clip_idx', 0))
    start = float(body['start']); end = float(body['end'])
    try:
        ssim_metric = _coerce_metric(body.get('ssim_metric'))
        ssim_threshold = _body_ssim_threshold(body, metric=ssim_metric)
    except ValueError as e:
        return jsonify({'status':'error','message':str(e)}),400
    raw_name = re.sub(r'[^\w\-]','_',(body.get('name') or '').strip())
    if not raw_name:
        return jsonify({'status':'error','message':'Nome não pode ser vazio'}),400

    occurrence = body.get('occurrence','sem_obstaculos')
    parts = [raw_name, occurrence]
    if occurrence == 'obstaculo':
        parts.append(body.get('obs_pos') or 'centro')
        response = body.get('response') or 'parada'
        parts.append(response)
        if response == 'desvio':
            parts.append(body.get('desvio_dir') or 'direita')

    folder_name = '_'.join(parts)

    # Collision check
    if name_exists(folder_name):
        return jsonify({'status':'error','message':f'Já existe um clipe com o nome "{folder_name}"'}),409

    export_folder = os.path.join(EXPORTS_DIR, folder_name)
    os.makedirs(export_folder)

    out_video = os.path.join(export_folder, folder_name+'.mp4')

    try:
        _ffmpeg_cut(clip_video_path(idx), start, end, out_video)
    except Exception as e:
        import shutil; shutil.rmtree(export_folder,ignore_errors=True)
        return jsonify({'status':'error','message':str(e)}),500

    # Save filtered sensor data alongside
    try:
        save_sensor_data(idx, start, end, export_folder)
    except Exception as e:
        pass  # sensor data save failure is non-fatal
    ssim_review = None
    try:
        selection = ssim_frame_selection(idx, start, end, threshold=ssim_threshold, metric=ssim_metric)
        save_ssim_selection(selection, export_folder)
        review_counts = save_ssim_review_videos(idx, start, end, selection, export_folder)
        ssim_review = {
            'counts': review_counts,
            'all_frames_url': f'/export-file/{folder_name}/ssim_review/all_frames.mp4',
            'chosen_frames_url': f'/export-file/{folder_name}/ssim_review/chosen_frames.mp4',
            'not_chosen_frames_url': f'/export-file/{folder_name}/ssim_review/not_chosen_frames.mp4',
        }
    except Exception:
        pass  # SSIM metadata is useful, but should not block the export

    # Update history
    try:
        number = int(raw_name) if raw_name.isdigit() else 0
    except:
        number = 0
    h = load_history()
    h.append({
        'number':       number,
        'folder':       folder_name,
        'source_clip':  CLIPS[idx],
        'source_idx':   idx,
        'start':        round(start, 3),
        'end':          round(end,   3),
        'duration':     round(end-start, 3),
        'occurrence':   occurrence,
        'obs_pos':      body.get('obs_pos'),
        'response':     body.get('response'),
        'desvio_dir':   body.get('desvio_dir'),
        'exported_at':  datetime.now().isoformat(timespec='seconds'),
    })
    save_history(h)

    return jsonify({'status':'ok','file':folder_name+'.mp4','folder':folder_name,
                    'export_folders': list_export_folders(),
                    'next_number': get_next_number(),
                    'ssim_review': ssim_review})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', '5000'))
    print(f'http://localhost:{port}  ({len(CLIPS)} clipes)')
    app.run(debug=False, host='0.0.0.0', port=port, threaded=True)
